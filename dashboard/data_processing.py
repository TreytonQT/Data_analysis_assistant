from __future__ import annotations

import io
import calendar
import math
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd

from dashboard.formula_engine import FormulaContext, FormulaError, evaluate_formula, extract_fields, extract_range_sums


ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = ROOT / "configs"

METRIC_COLUMNS = ["指标名称", "显示分组", "公式", "格式", "排序", "是否启用"]
STORE_COLUMNS = ["店铺名", "店铺类型", "停提款时间", "店铺所属部门"]
TARGET_COLUMNS = ["开发员", "目标业绩", "目标毛利率"]
COMMISSION_COLUMNS = ["月份", "开发员", "库存计提", "弃置", "职位提点"]
DEPARTMENT_FEE_COLUMNS = ["月份", "部门", "费用率"]
# Deprecated compatibility constants.  The replenishment page now gets its
# target coverage from rules instead of an ASIN-level target/case-pack table.
REPLENISHMENT_TARGET_COLUMNS = ["ASIN", "目标可售天数"]
DEFAULT_REPLENISHMENT_TARGET_DAYS = 70
DEFAULT_REPLENISHMENT_CASE_PACK = 10
REPLENISHMENT_COVERAGE_RULE_COLUMNS = [
    "运输方式", "重量下限", "重量上限", "头程时效", "预警天数", "补货频次", "是否启用",
]
REPLENISHMENT_SWITCH_COLUMNS = ["ASIN", "是否补货", "关闭原因"]
REPLENISHMENT_PRODUCT_TAG_COLUMNS = ["ASIN", "产品标签", "标签颜色", "是否启用", "备注"]
SALES_HISTORY_2025_SHEETS = [f"{month}月" for month in range(1, 13)]
SALES_HISTORY_2025_COUNTRIES = {
    "德国": "DE",
    "法国": "FR",
    "西班牙": "ES",
    "意大利": "IT",
}
SALES_HISTORY_2025_SITE_COLUMNS = [f"{code}总销量" for code in SALES_HISTORY_2025_COUNTRIES.values()]
SALES_HISTORY_2025_MONTH_COLUMNS = [
    column
    for month in range(1, 13)
    for column in (f"{month}月总销量", f"{month}月出单天数", f"{month}月除0日均")
]
SALES_HISTORY_2025_COLUMNS = ["ASIN", *SALES_HISTORY_2025_SITE_COLUMNS, *SALES_HISTORY_2025_MONTH_COLUMNS]
SALES_HISTORY_GENERIC_MONTH_COLUMNS = [
    column
    for index in range(1, 13)
    for column in (
        f"历史月份{index}",
        f"历史{index}月总销量",
        f"历史{index}月计入天数",
        f"历史{index}月日均销量",
    )
]
SALES_HISTORY_GENERIC_COLUMNS = ["ASIN", *SALES_HISTORY_2025_SITE_COLUMNS, *SALES_HISTORY_GENERIC_MONTH_COLUMNS]
OPERATIONAL_SALES_REQUIRED_COLUMNS = [
    "MSKU",
    "店铺名称",
    "7天销量",
    "30天销量",
    "可售",
    "本地库存",
    "昨天销量",
    "前天销量",
    "上前销量",
    "开发员",
    "ASIN",
]
OPERATIONAL_SALES_NUMERIC_COLUMNS = [
    "7天销量",
    "30天销量",
    "可售",
    "本地库存",
    "昨天销量",
    "前天销量",
    "上前销量",
]
OPERATIONAL_SALES_DERIVED_COLUMNS = ["占用资金"]
OPERATIONAL_SALES_NORMALIZED_COLUMNS = OPERATIONAL_SALES_REQUIRED_COLUMNS + OPERATIONAL_SALES_DERIVED_COLUMNS + [
    "店铺编码",
    "店铺名称原始",
    "店铺名称展开",
    "店铺类型推断",
    "是否多店铺编码",
    "30天日均",
    "7天日均",
    "是否在售",
    "是否-26",
]
AGING_STOCK_COLUMNS = [
    "91-180天库存数",
    "181-330天库存数",
    "331-365天库存数",
    "366-455天库存数",
    "456天以上库存数",
]
AGING_CAPITAL_COLUMNS = [
    "91-180天占用资金",
    "181-330天占用资金",
    "331-365天占用资金",
    "366-455天占用资金",
    "456天占用资金",
]
OPERATIONAL_AGING_REQUIRED_COLUMNS = ["MSKU", "开发员", "ASIN"] + AGING_STOCK_COLUMNS + AGING_CAPITAL_COLUMNS
REPLENISHMENT_STOCK_COMPONENT_COLUMNS = ["可售", "待入库", "采购在途", "本地库存", "在途", "计划入库"]
REPLENISHMENT_FORMULA_STOCK_COLUMNS = [
    "可售", "待调仓", "调仓中", "待入库", "采购在途", "本地库存", "在途", "计划入库",
]
REPLENISHMENT_SALES_COLUMNS = ["7天销量", "14天销量", "30天销量"]
AVAILABLE_INVENTORY_STOCK_COLUMNS = [
    "可售",
    "待调仓",
    "调仓中",
    "待入库",
    "采购在途",
    "本地库存",
    "在途",
    "计划入库",
]
AVAILABLE_INVENTORY_REQUIRED_COLUMNS = ["开发员", "日均销量"] + AVAILABLE_INVENTORY_STOCK_COLUMNS
AVAILABLE_INVENTORY_SUMMARY_COLUMNS = ["开发员"] + AVAILABLE_INVENTORY_STOCK_COLUMNS + [
    "库存总数",
    "日均单量",
    "总可售天数",
]
AVAILABLE_INVENTORY_MONITOR_COLUMNS = ["开发员", "库存总数", "日均订单", "总可售天数"]
SALES_VOLUME_DETAIL_REQUIRED_COLUMNS = ["msku", "店铺", "开发专员"]
SALES_AMOUNT_DETAIL_REQUIRED_COLUMNS = ["msku", "店铺", "开发专员"]
DEPARTMENT_PERFORMANCE_FIXED_COLUMNS = [
    "在售SKU数量",
    "销售额贡献占比",
    "近7天日均订单",
    "近7天日均销售额（元）",
    "预估本月销售额（元）",
]
REPLENISHMENT_OPERATIONAL_REQUIRED_COLUMNS = [
    "ASIN",
    "MSKU",
    "店铺名称",
    "开发员",
    "单品重量(g)",
    "上架时间",
] + REPLENISHMENT_SALES_COLUMNS + REPLENISHMENT_FORMULA_STOCK_COLUMNS
DISCARD_THRESHOLD_SEGMENTS = {
    "90天以上": ["91-180", "181-330", "331-365", "366-455", "456天以上"],
    "180天以上": ["181-330", "331-365", "366-455", "456天以上"],
    "365天以上": ["366-455", "456天以上"],
}
PRODUCT_COUNTRIES = ["德国", "法国", "西班牙", "意大利"]
PRODUCT_OPERATIONAL_REQUIRED_COLUMNS = [
    "ASIN",
    "MSKU",
    "可售",
    "可售天数",
    "日均销量",
    "昨天销量",
    "前天销量",
    "上前销量",
    "7天销量",
    "14天销量",
    "30天销量",
    "90天销量",
]
PRODUCT_OPERATIONAL_SUM_COLUMNS = [
    "可售数量",
    "日均销量",
    "昨天销量",
    "前天销量",
    "上前销量",
    "7天销量",
    "14天销量",
    "30天销量",
    "90天销量",
]
GROSS_PROFIT_VOLUME_COLUMNS = ["销量--FBA销量", "销量--FBM销量", "销量--多渠道销量"]
REPLENISHMENT_GROSS_RATIO_COLUMNS = ["广告费占比", "退款占比", "FBA发货费占比"]
REPLENISHMENT_GROSS_REQUIRED_COLUMNS = [
    "ASIN",
    "MSKU",
    "国家",
    "销售额--FBA销售额",
    "COD",
    "毛利润",
] + GROSS_PROFIT_VOLUME_COLUMNS + REPLENISHMENT_GROSS_RATIO_COLUMNS
GROSS_PROFIT_AD_COLUMNS = [
    "广告费-SD广告",
    "广告费-SP广告",
    "广告费-SB广告",
    "广告费-SBV广告",
    "广告费--差异分摊",
]
GROSS_PROFIT_REQUIRED_COLUMNS = [
    "ASIN",
    "MSKU",
    "国家",
    "销售额--FBA销售额",
    "COD",
    "毛利润",
] + GROSS_PROFIT_VOLUME_COLUMNS + GROSS_PROFIT_AD_COLUMNS
RATING_REQUIRED_COLUMNS = ["ASIN", "国家", "Rating总数", "评分"]
PRODUCT_LEVELS = [
    ("0单", lambda value: value == 0),
    ("0.2单以下", lambda value: 0 < value <= 0.2),
    ("0.2-0.5单", lambda value: 0.2 < value <= 0.5),
    ("0.5-1单", lambda value: 0.5 < value <= 1),
    ("1-2单", lambda value: 1 < value <= 2),
    ("2-3单", lambda value: 2 < value <= 3),
    ("3-5单", lambda value: 3 < value <= 5),
    ("5单以上", lambda value: value > 5),
]
LOW_MARGIN_PRODUCT_THRESHOLD = 0.15
LOW_MARGIN_PRODUCT_MIN_SALES = 5
LOW_MARGIN_PRODUCT_COLUMNS = ["SKU", "ASIN", "国家", "开发员", "销量", "销售额", "毛利润", "毛利率"]
GROSS_PROFIT_DEVELOPER_COLUMNS = ["开发员", "开发人员", "销售专员", "销售"]
COMMISSION_OUTPUT_COLUMNS = [
    "月份",
    "开发员",
    "营业额",
    "毛利润",
    "毛利率",
    "费用率",
    "库存计提",
    "弃置",
    "职位提点",
    "提成预估",
    "配置状态",
]
STOPPED_COMMISSION_OUTPUT_COLUMNS = [
    "月份",
    "开发员",
    "店铺编码",
    "店铺类型",
    "部门",
    "停提款时间",
    "营业额",
    "毛利润",
    "毛利率",
    "费用率",
    "库存计提分摊",
    "弃置分摊",
    "职位提点",
    "缺提成预估",
    "配置状态",
]


def read_upload_table(uploaded_file, fallback_path: Path | None = None) -> pd.DataFrame:
    if uploaded_file is None:
        if fallback_path is None or not fallback_path.exists():
            return pd.DataFrame()
        return read_local_table(fallback_path)

    name = uploaded_file.name.lower()
    data = uploaded_file.getvalue()
    if name.endswith(".xlsx"):
        return pd.read_excel(io.BytesIO(data))
    if name.endswith(".xls"):
        return pd.read_excel(io.BytesIO(data), engine="xlrd", engine_kwargs={"ignore_workbook_corruption": True})
    return read_csv_bytes(data)


def read_local_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".xlsx":
        return pd.read_excel(path)
    if path.suffix.lower() == ".xls":
        return pd.read_excel(path, engine="xlrd", engine_kwargs={"ignore_workbook_corruption": True})
    return read_csv_bytes(path.read_bytes())


def read_csv_bytes(data: bytes) -> pd.DataFrame:
    last_error = None
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            return pd.read_csv(io.BytesIO(data), encoding=encoding, low_memory=False)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"CSV 编码无法识别：{last_error}")


def load_metric_config(uploaded_file=None) -> pd.DataFrame:
    df = read_upload_table(uploaded_file, CONFIG_DIR / "metrics_config.csv")
    missing = [col for col in METRIC_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"指标配置缺少列：{', '.join(missing)}")

    df = df[METRIC_COLUMNS].copy()
    df["排序"] = pd.to_numeric(df["排序"], errors="coerce").fillna(9999)
    df["是否启用"] = df["是否启用"].map(is_enabled)
    df = df[df["是否启用"]].sort_values(["显示分组", "排序", "指标名称"]).reset_index(drop=True)
    if df.empty:
        raise ValueError("没有启用的指标")
    return df


def load_business_config(
    store_config: pd.DataFrame | None = None, target_config: pd.DataFrame | None = None
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if store_config is None:
        store_config = read_upload_table(None, CONFIG_DIR / "store_config.csv")
    if target_config is None:
        target_config = read_upload_table(None, CONFIG_DIR / "monthly_targets.csv")

    return normalize_store_config(store_config), normalize_target_config(target_config)


def load_commission_config(uploaded_file=None) -> pd.DataFrame:
    return normalize_commission_config(read_upload_table(uploaded_file, CONFIG_DIR / "commission_config.csv"))


def load_department_fee_config(uploaded_file=None) -> pd.DataFrame:
    return normalize_department_fee_config(read_upload_table(uploaded_file, CONFIG_DIR / "department_fee_config.csv"))


def load_operational_sales_source(path_or_file) -> pd.DataFrame:
    return normalize_operational_sales(read_upload_table(path_or_file) if hasattr(path_or_file, "getvalue") else read_local_table(Path(path_or_file)))


def normalize_store_config(store_config: pd.DataFrame) -> pd.DataFrame:
    store_config = store_config.copy()
    aliases = {
        "店铺编码": "店铺名",
        "部门": "店铺所属部门",
        "停提款月份": "停提款时间",
    }
    store_config = store_config.rename(columns={old: new for old, new in aliases.items() if old in store_config.columns})

    if store_config.empty:
        store_config = pd.DataFrame(columns=STORE_COLUMNS)
    for col in STORE_COLUMNS:
        if col not in store_config.columns:
            store_config[col] = None
    store_config = store_config[STORE_COLUMNS].copy()
    store_config["停提款时间"] = store_config["停提款时间"].map(normalize_config_month).fillna("")
    store_config = store_config[store_config["店铺名"].notna()].drop_duplicates(subset=["店铺名"], keep="first")
    return store_config


def normalize_target_config(target_config: pd.DataFrame) -> pd.DataFrame:
    target_config = target_config.copy()
    aliases = {
        "销售专员": "开发员",
        "销售额目标": "目标业绩",
        "销售目标": "目标业绩",
        "毛利率目标": "目标毛利率",
        "毛利率": "目标毛利率",
    }
    target_config = target_config.rename(columns={old: new for old, new in aliases.items() if old in target_config.columns})
    if target_config.empty:
        target_config = pd.DataFrame(columns=TARGET_COLUMNS)
    for col in TARGET_COLUMNS:
        if col not in target_config.columns:
            target_config[col] = None
    target_config = target_config[TARGET_COLUMNS].copy()
    target_config = target_config[target_config["开发员"].notna()].drop_duplicates(subset=["开发员"], keep="first")
    return target_config


def normalize_commission_config(commission_config: pd.DataFrame) -> pd.DataFrame:
    commission_config = commission_config.copy()
    aliases = {
        "销售专员": "开发员",
        "月份": "月份",
        "库存": "库存计提",
        "提点": "职位提点",
        "职位提成点": "职位提点",
    }
    commission_config = commission_config.rename(
        columns={old: new for old, new in aliases.items() if old in commission_config.columns}
    )
    if commission_config.empty:
        commission_config = pd.DataFrame(columns=COMMISSION_COLUMNS)
    for col in COMMISSION_COLUMNS:
        if col not in commission_config.columns:
            commission_config[col] = None
    commission_config = commission_config[COMMISSION_COLUMNS].copy()
    commission_config["月份"] = commission_config["月份"].map(normalize_month)
    commission_config = commission_config[
        commission_config["月份"].notna() & commission_config["开发员"].notna()
    ].drop_duplicates(subset=["月份", "开发员"], keep="first")
    commission_config["职位提点"] = normalize_rate(commission_config["职位提点"])
    commission_config["库存计提"] = normalize_config_number(commission_config["库存计提"])
    commission_config["弃置"] = normalize_config_number(commission_config["弃置"])
    return commission_config.reset_index(drop=True)


def normalize_department_fee_config(department_fee_config: pd.DataFrame) -> pd.DataFrame:
    department_fee_config = department_fee_config.copy()
    aliases = {
        "店铺所属部门": "部门",
        "费用部门": "部门",
        "费率": "费用率",
    }
    department_fee_config = department_fee_config.rename(
        columns={old: new for old, new in aliases.items() if old in department_fee_config.columns}
    )
    if department_fee_config.empty:
        department_fee_config = pd.DataFrame(columns=DEPARTMENT_FEE_COLUMNS)
    for col in DEPARTMENT_FEE_COLUMNS:
        if col not in department_fee_config.columns:
            department_fee_config[col] = None
    department_fee_config = department_fee_config[DEPARTMENT_FEE_COLUMNS].copy()
    department_fee_config["月份"] = department_fee_config["月份"].map(normalize_month)
    department_fee_config = department_fee_config[
        department_fee_config["月份"].notna() & department_fee_config["部门"].notna()
    ].drop_duplicates(subset=["月份", "部门"], keep="first")
    department_fee_config["费用率"] = normalize_rate(department_fee_config["费用率"])
    return department_fee_config.reset_index(drop=True)


def load_reports(files: Iterable) -> pd.DataFrame:
    frames = []
    for file in files:
        if hasattr(file, "getvalue"):
            frame = read_csv_bytes(file.getvalue()).copy()
            frame["来源文件"] = file.name
        else:
            path = Path(file)
            frame = read_local_table(path).copy()
            frame["来源文件"] = path.name
        frames.append(frame)
    if not frames:
        raise ValueError("没有上传业绩报表")
    df = pd.concat(frames, ignore_index=True)
    required = ["销售专员", "月份", "店铺"]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"业绩报表缺少基础列：{', '.join(missing)}")
    return normalize_report(df)


def normalize_report(df: pd.DataFrame, *, today: date | None = None) -> pd.DataFrame:
    result = df.copy()
    result["月份"] = result["月份"].map(lambda value: normalize_month(value, today=today))
    result["店铺编码"] = result["店铺"].map(extract_store_code)
    for col in result.columns:
        if col in {"销售专员", "月份", "国家", "店铺", "店铺编码", "来源文件"}:
            continue
        result[col] = maybe_numeric(result[col])
    return result


def normalize_month(value, *, today: date | None = None) -> str | None:
    if pd.isna(value):
        return None
    text = str(value)
    text = text.strip()
    if not text:
        return None
    range_match = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})\s*[~～]\s*(\d{4})-(\d{2})-(\d{2})", text)
    if range_match:
        start_year, start_month, start_day, end_year, end_month, end_day = map(int, range_match.groups())
        try:
            start = date(start_year, start_month, start_day)
            end = date(end_year, end_month, end_day)
        except ValueError:
            return None
        if (start.year, start.month) != (end.year, end.month) or start > end:
            return None
        current_day = today or date.today()
        is_current_month = (start.year, start.month) == (current_day.year, current_day.month)
        last_day = calendar.monthrange(start.year, start.month)[1]
        if is_current_month or (start.day == 1 and end.day == last_day):
            return f"{start_year:04d}-{start_month:02d}"
        return None
    chinese_match = re.fullmatch(r"(\d{2,4})\s*年\s*(\d{1,2})\s*月", text)
    if chinese_match:
        year = int(chinese_match.group(1))
        if year < 100:
            year += 2000
        month = int(chinese_match.group(2))
        return f"{year:04d}-{month:02d}" if 1 <= month <= 12 else None
    match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", text)
    if not match:
        return None
    month = int(match.group(2))
    return f"{int(match.group(1)):04d}-{month:02d}" if 1 <= month <= 12 else None


def normalize_config_month(value) -> str | None:
    normalized = normalize_month(value)
    if normalized is None:
        return None
    return normalized if re.fullmatch(r"\d{4}-\d{2}", str(normalized)) else None


def extract_store_code(value) -> str:
    if pd.isna(value):
        return "未识别"
    text = str(value).strip()
    match = re.search(r"^[^-]+-([A-Za-z0-9]+)", text)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b([A-Za-z]{2,5})\b", text)
    return match.group(1).upper() if match else text


def extract_operational_store_codes(value) -> list[tuple[str, str]]:
    if pd.isna(value):
        return [("未识别", "")]
    text = str(value).strip()
    if not text:
        return [("未识别", "")]

    stores = []
    seen = set()
    for item in [part.strip() for part in text.split(",") if part.strip()]:
        code = extract_store_code(item)
        if code not in seen:
            stores.append((code, item))
            seen.add(code)
    return stores or [("未识别", text)]


def infer_operational_store_type(store_name: str) -> str:
    return "本土" if "本土" in str(store_name) else "中企"


def normalize_operational_sales(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in OPERATIONAL_SALES_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"运营原始表缺少列：{', '.join(missing)}")

    base = df[OPERATIONAL_SALES_REQUIRED_COLUMNS].copy()
    capital_cols = operational_sales_capital_columns(df)
    for col in capital_cols:
        base[col] = normalize_config_number(df[col]).fillna(0)
    base["占用资金"] = base[capital_cols].sum(axis=1) if capital_cols else 0
    for col in ["MSKU", "店铺名称", "开发员", "ASIN"]:
        base[col] = base[col].fillna("").astype(str).str.strip()
    for col in OPERATIONAL_SALES_NUMERIC_COLUMNS:
        base[col] = normalize_config_number(base[col]).fillna(0)

    rows = []
    for _, row in base.iterrows():
        stores = extract_operational_store_codes(row["店铺名称"])
        is_multi_store_code = len(stores) > 1
        for store_code, store_name in stores:
            item = row.to_dict()
            item["店铺编码"] = store_code
            item["店铺名称原始"] = row["店铺名称"]
            item["店铺名称展开"] = store_name
            item["店铺类型推断"] = infer_operational_store_type(store_name)
            item["是否多店铺编码"] = is_multi_store_code
            item["30天日均"] = item["30天销量"] / 30
            item["7天日均"] = item["7天销量"] / 7
            item["是否在售"] = item["可售"] > 0
            item["是否-26"] = str(item["开发员"]).strip().endswith("-26")
            rows.append(item)
    if not rows:
        return pd.DataFrame(columns=OPERATIONAL_SALES_NORMALIZED_COLUMNS)
    return pd.DataFrame(rows)


def operational_sales_capital_columns(df: pd.DataFrame) -> list[str]:
    return [
        col
        for col in df.columns
        if re.fullmatch(r"\d+(?:-\d+)?天(?:以上)?占用资金", str(col).strip())
    ]


def ensure_operational_sales_normalized(df: pd.DataFrame) -> pd.DataFrame:
    if all(col in df.columns for col in OPERATIONAL_SALES_NORMALIZED_COLUMNS):
        return df.copy()
    return normalize_operational_sales(df)


def count_chen_26_onsale_skus(df: pd.DataFrame) -> int:
    data = ensure_operational_sales_normalized(df)
    if data.empty:
        return 0
    developer = data["开发员"].fillna("").astype(str).str.strip()
    onsale = normalize_config_number(data["可售"]).fillna(0).gt(0)
    matched = data.loc[developer.str.endswith("陈千潼-26") & onsale, "MSKU"].fillna("").astype(str).str.strip()
    return int(matched[matched.ne("")].nunique())


def merge_operational_store_config(df: pd.DataFrame, store_config: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    config = normalize_store_config(store_config).copy()
    if not config.empty:
        config["店铺编码"] = config["店铺名"].map(extract_store_code)
        config = config.drop_duplicates(subset=["店铺编码"], keep="first")
        result = result.merge(config[["店铺编码", "店铺类型", "停提款时间"]], on="店铺编码", how="left")
    for col in ["店铺类型", "停提款时间"]:
        if col not in result.columns:
            result[col] = pd.NA
    result["店铺类型"] = result["店铺类型"].where(result["店铺类型"].notna() & result["店铺类型"].astype(str).str.strip().ne(""), result["店铺类型推断"])
    result["停提款时间"] = result["停提款时间"].fillna("").astype(str).str.strip()
    result["是否封店"] = result["停提款时间"].ne("")
    result["店铺状态"] = result["是否封店"].map(lambda value: "已封店" if value else "正常")
    return result


def exclude_stopped_store_operational_rows(
    df: pd.DataFrame, store_config: pd.DataFrame | None
) -> pd.DataFrame:
    """Remove raw operational rows belonging to stores configured with a stop month.

    Slow-moving inventory and promotion candidates are calculated from the raw
    operational file, before the normal store-expansion step.  A configured
    stop month means the store is no longer actionable, regardless of the
    specific month value, so its rows must not contribute inventory or sales to
    these operational reminders.
    """
    if df.empty or "店铺名称" not in df.columns or store_config is None:
        return df.copy()

    config = normalize_store_config(store_config)
    stop_month = config["停提款时间"].fillna("").astype(str).str.strip()
    stopped_codes = {
        extract_store_code(store_name)
        for store_name in config.loc[stop_month.ne(""), "店铺名"]
        if extract_store_code(store_name)
    }
    if not stopped_codes:
        return df.copy()

    is_stopped = df["店铺名称"].map(
        lambda value: any(code in stopped_codes for code, _ in extract_operational_store_codes(value))
    )
    return df.loc[~is_stopped].copy()


def safe_ratio(numerator, denominator):
    return numerator / denominator if denominator else 0


def product_level_for_daily_sales(value: float) -> str:
    for label, predicate in PRODUCT_LEVELS:
        if predicate(value):
            return label
    return "未分档"


def build_sales_dashboard_tables(df: pd.DataFrame, store_config: pd.DataFrame) -> dict[str, pd.DataFrame]:
    data = merge_operational_store_config(ensure_operational_sales_normalized(df), store_config)
    if data.empty:
        empty = pd.DataFrame()
        return {"stores": empty, "levels": empty, "date_compare": empty, "type_compare": empty, "source": data}

    store_order = []
    if store_config is not None and not store_config.empty:
        store_order = normalize_store_config(store_config)["店铺名"].map(extract_store_code).dropna().drop_duplicates().tolist()

    data["产品等级"] = data["30天日均"].map(product_level_for_daily_sales)
    data["有效在售"] = data["是否在售"] & ~data["是否封店"]
    total_onsale = data["有效在售"].sum()
    total_30_avg = data["30天日均"].sum()

    store_summary = (
        data.groupby(["店铺编码", "店铺类型", "店铺状态"], dropna=False, as_index=False)
        .agg(
            在售个数=("有效在售", "sum"),
            昨日订单=("昨天销量", "sum"),
            前天订单=("前天销量", "sum"),
            上前订单=("上前销量", "sum"),
            **{"-26订单": ("昨天销量", lambda values: values[data.loc[values.index, "是否-26"]].sum())},
            **{"7天日均": ("7天日均", "sum")},
            **{"30天日均": ("30天日均", "sum")},
            总库存=("可售", "sum"),
            占用资金=("占用资金", "sum"),
        )
    )
    store_summary["产品数占比"] = store_summary["在售个数"].map(lambda value: safe_ratio(value, total_onsale))
    store_summary["昨日D值"] = store_summary.apply(lambda row: safe_ratio(row["昨日订单"], row["在售个数"]), axis=1)
    store_summary["7天D值"] = store_summary.apply(lambda row: safe_ratio(row["7天日均"], row["在售个数"]), axis=1)
    order_lookup = {code: idx for idx, code in enumerate(store_order)}
    store_summary["_排序"] = store_summary["店铺编码"].map(order_lookup).fillna(len(order_lookup) + 999)
    store_summary = store_summary.sort_values(["_排序", "店铺编码"]).drop(columns=["_排序"]).reset_index(drop=True)
    store_summary = store_summary[
        [
            "店铺编码",
            "店铺类型",
            "店铺状态",
            "在售个数",
            "产品数占比",
            "昨日D值",
            "7天D值",
            "昨日订单",
            "-26订单",
            "7天日均",
            "30天日均",
            "总库存",
            "占用资金",
        ]
    ]

    level_summary = (
        data.groupby("产品等级", dropna=False, as_index=False)
        .agg(
            在售个数=("有效在售", "sum"),
            昨日订单=("昨天销量", "sum"),
            **{"7天日均": ("7天日均", "sum")},
            **{"30天日均": ("30天日均", "sum")},
        )
    )
    level_summary = pd.DataFrame({"产品等级": [label for label, _ in PRODUCT_LEVELS]}).merge(
        level_summary, on="产品等级", how="left"
    )
    for col in ["在售个数", "昨日订单", "7天日均", "30天日均"]:
        level_summary[col] = level_summary[col].fillna(0)
    level_order = {label: idx for idx, (label, _) in enumerate(PRODUCT_LEVELS)}
    level_summary["_排序"] = level_summary["产品等级"].map(level_order).fillna(999)
    level_summary["产品数占比"] = level_summary["在售个数"].map(lambda value: safe_ratio(value, total_onsale))
    level_summary["30天贡献占比"] = level_summary["30天日均"].map(lambda value: safe_ratio(value, total_30_avg))
    level_summary = level_summary.sort_values("_排序").drop(columns=["_排序"]).reset_index(drop=True)
    total_row = {
        "产品等级": "总计",
        "在售个数": level_summary["在售个数"].sum(),
        "产品数占比": safe_ratio(level_summary["在售个数"].sum(), total_onsale),
        "昨日订单": level_summary["昨日订单"].sum(),
        "7天日均": level_summary["7天日均"].sum(),
        "30天日均": level_summary["30天日均"].sum(),
        "30天贡献占比": safe_ratio(level_summary["30天日均"].sum(), total_30_avg),
    }
    level_summary = pd.concat([level_summary, pd.DataFrame([total_row])], ignore_index=True)
    level_summary = level_summary[["产品等级", "在售个数", "产品数占比", "昨日订单", "7天日均", "30天日均", "30天贡献占比"]]

    type_compare = (
        data.groupby("店铺类型", dropna=False, as_index=False)
        .agg(
            昨天=("昨天销量", "sum"),
            前天=("前天销量", "sum"),
            上前=("上前销量", "sum"),
            **{"7天": ("7天日均", "sum")},
            **{"30天": ("30天日均", "sum")},
        )
        .set_index("店铺类型")
    )
    date_compare = pd.DataFrame(
        [
            {"日期": "昨天", "中企单量": type_compare["昨天"].get("中企", 0), "本土单量": type_compare["昨天"].get("本土", 0)},
            {"日期": "前天", "中企单量": type_compare["前天"].get("中企", 0), "本土单量": type_compare["前天"].get("本土", 0)},
            {"日期": "上前", "中企单量": type_compare["上前"].get("中企", 0), "本土单量": type_compare["上前"].get("本土", 0)},
            {"日期": "7天", "中企单量": type_compare["7天"].get("中企", 0), "本土单量": type_compare["7天"].get("本土", 0)},
            {"日期": "30天", "中企单量": type_compare["30天"].get("中企", 0), "本土单量": type_compare["30天"].get("本土", 0)},
        ]
    )
    date_compare["总计"] = date_compare["中企单量"] + date_compare["本土单量"]

    return {
        "stores": store_summary,
        "levels": level_summary,
        "date_compare": date_compare,
        "source": data,
    }


def normalize_operational_aging(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in OPERATIONAL_AGING_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"运营原始表缺少库龄列：{', '.join(missing)}")

    result = df[OPERATIONAL_AGING_REQUIRED_COLUMNS].copy()
    for col in ["MSKU", "开发员", "ASIN"]:
        result[col] = result[col].fillna("").astype(str).str.strip()
    for col in AGING_STOCK_COLUMNS + AGING_CAPITAL_COLUMNS:
        result[col] = normalize_config_number(result[col]).fillna(0)
    return result


def build_slow_moving_inventory_table(df: pd.DataFrame, discard_threshold: str = "90天以上") -> pd.DataFrame:
    if discard_threshold not in DISCARD_THRESHOLD_SEGMENTS:
        raise ValueError(f"未知弃置费阈值：{discard_threshold}")

    data = normalize_operational_aging(df)
    if data.empty:
        return pd.DataFrame()

    agg = {col: (col, "sum") for col in AGING_STOCK_COLUMNS + AGING_CAPITAL_COLUMNS}
    base = data.groupby("MSKU", dropna=False, as_index=False).agg(
        开发员=("开发员", lambda values: "；".join(sorted({str(value) for value in values if str(value).strip()}))),
        ASIN=("ASIN", lambda values: "；".join(sorted({str(value) for value in values if str(value).strip()}))),
        **agg,
    )

    stock_90_plus = base[AGING_STOCK_COLUMNS].sum(axis=1)
    capital_90_plus = base[AGING_CAPITAL_COLUMNS].sum(axis=1)
    base = base[stock_90_plus.gt(0)].copy()
    if base.empty:
        return pd.DataFrame(columns=slow_moving_inventory_columns())

    base["90天以上库存数合计"] = base[AGING_STOCK_COLUMNS].sum(axis=1)
    base["90天以上占用资金合计"] = base[AGING_CAPITAL_COLUMNS].sum(axis=1)
    base["库存计提"] = (
        base["91-180天占用资金"] * 0.05
        + (base["181-330天占用资金"] + base["331-365天占用资金"]) * 0.08
        + (base["366-455天占用资金"] + base["456天占用资金"]) * 0.12
    )

    segment_to_stock = {
        "91-180": "91-180天库存数",
        "181-330": "181-330天库存数",
        "331-365": "331-365天库存数",
        "366-455": "366-455天库存数",
        "456天以上": "456天以上库存数",
    }
    segment_to_capital = {
        "91-180": "91-180天占用资金",
        "181-330": "181-330天占用资金",
        "331-365": "331-365天占用资金",
        "366-455": "366-455天占用资金",
        "456天以上": "456天占用资金",
    }
    segments = DISCARD_THRESHOLD_SEGMENTS[discard_threshold]
    discard_stock = base[[segment_to_stock[segment] for segment in segments]].sum(axis=1)
    discard_capital = base[[segment_to_capital[segment] for segment in segments]].sum(axis=1)
    base["弃置费"] = discard_stock * 6 + discard_capital * 1.5

    base = base.rename(columns={"MSKU": "SKU"})
    return base[slow_moving_inventory_columns()].sort_values("90天以上占用资金合计", ascending=False).reset_index(drop=True)


def slow_moving_inventory_columns() -> list[str]:
    return [
        "SKU",
        "开发员",
        "ASIN",
        "90天以上库存数合计",
        "90天以上占用资金合计",
        "库存计提",
        "弃置费",
        "91-180天库存数",
        "181-330天库存数",
        "331-365天库存数",
        "366-455天库存数",
        "456天以上库存数",
        "91-180天占用资金",
        "181-330天占用资金",
        "331-365天占用资金",
        "366-455天占用资金",
        "456天占用资金",
    ]


def normalize_available_inventory_monitor(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in AVAILABLE_INVENTORY_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"运营原始表缺少可售监控列：{', '.join(missing)}")

    base = df[AVAILABLE_INVENTORY_REQUIRED_COLUMNS].copy()
    base["开发员"] = base["开发员"].fillna("").astype(str).str.strip()
    base = base[base["开发员"].ne("")].copy()
    for col in ["日均销量"] + AVAILABLE_INVENTORY_STOCK_COLUMNS:
        base[col] = normalize_config_number(base[col]).fillna(0)

    if base.empty:
        return pd.DataFrame(columns=AVAILABLE_INVENTORY_SUMMARY_COLUMNS)

    grouped = (
        base.groupby("开发员", dropna=False, sort=True)[["日均销量"] + AVAILABLE_INVENTORY_STOCK_COLUMNS]
        .sum()
        .reset_index()
    )
    grouped["库存总数"] = grouped[AVAILABLE_INVENTORY_STOCK_COLUMNS].sum(axis=1)
    grouped["日均单量"] = grouped["日均销量"]
    grouped["总可售天数"] = grouped.apply(lambda row: safe_blank_ratio(row["库存总数"], row["日均单量"]), axis=1)
    return grouped[AVAILABLE_INVENTORY_SUMMARY_COLUMNS].reset_index(drop=True)


def build_available_inventory_monitor_table(df: pd.DataFrame) -> pd.DataFrame:
    summary = normalize_available_inventory_monitor(df)
    if summary.empty:
        return pd.DataFrame(columns=AVAILABLE_INVENTORY_MONITOR_COLUMNS)

    result = summary[["开发员", "库存总数", "日均单量", "总可售天数"]].rename(columns={"日均单量": "日均订单"})
    return result[AVAILABLE_INVENTORY_MONITOR_COLUMNS].reset_index(drop=True)


def sales_detail_date_columns(df: pd.DataFrame, suffix: str) -> list[str]:
    pattern = re.compile(r"^\d{2}-\d{2}" + re.escape(suffix) + r"$")
    return [col for col in df.columns if pattern.fullmatch(str(col).strip())]


def normalize_sales_volume_detail(df: pd.DataFrame) -> pd.DataFrame:
    return normalize_sales_metric_detail(df, SALES_VOLUME_DETAIL_REQUIRED_COLUMNS, "销量")


def normalize_sales_amount_detail(df: pd.DataFrame) -> pd.DataFrame:
    return normalize_sales_metric_detail(df, SALES_AMOUNT_DETAIL_REQUIRED_COLUMNS, "销售额")


def duplicate_row_issues(frame: pd.DataFrame, example_limit: int = 5) -> list[dict]:
    """Describe every group of completely duplicate rows for upload validation.

    ``example_limit`` limits reported row numbers per group, not the number of
    groups, so summing ``duplicate_count`` always yields the exact total.
    """

    if not isinstance(frame, pd.DataFrame):
        raise TypeError("重复行检测只接受 DataFrame")
    if example_limit < 1:
        raise ValueError("每组示例行数必须大于 0")
    if frame.empty or not frame.duplicated(keep=False).any():
        return []

    def value_key(value):
        try:
            missing = pd.isna(value)
            if isinstance(missing, bool) and missing:
                return ("missing",)
        except (TypeError, ValueError):
            pass
        try:
            hash(value)
            return (type(value).__name__, value)
        except TypeError:
            return (type(value).__name__, repr(value))

    groups: dict[tuple, list[int]] = {}
    examples: dict[tuple, tuple] = {}
    for row_number, values in enumerate(frame.itertuples(index=False, name=None), start=1):
        key = tuple(value_key(value) for value in values)
        groups.setdefault(key, []).append(row_number)
        examples.setdefault(key, values)

    issues = []
    for key, row_numbers in groups.items():
        if len(row_numbers) < 2:
            continue
        values = examples[key]
        example = {}
        for column, value in zip(frame.columns, values):
            try:
                is_missing = bool(pd.isna(value))
            except (TypeError, ValueError):
                is_missing = False
            if is_missing:
                example[str(column)] = None
            elif hasattr(value, "item"):
                example[str(column)] = value.item()
            else:
                example[str(column)] = value
        issues.append(
            {
                "duplicate_count": len(row_numbers) - 1,
                "row_numbers": row_numbers[:example_limit],
                "example": example,
            }
        )
    return issues


def normalize_sales_metric_detail(df: pd.DataFrame, required_columns: list[str], suffix: str) -> pd.DataFrame:
    missing = [col for col in required_columns if col not in df.columns]
    date_columns = sales_detail_date_columns(df, suffix)
    if missing:
        raise ValueError(f"{suffix}明细缺少列：{', '.join(missing)}")
    if not date_columns:
        raise ValueError(f"{suffix}明细缺少 MM-DD{suffix} 日期列")

    result = df[required_columns + date_columns].copy()
    for col in required_columns:
        result[col] = result[col].fillna("").astype(str).str.strip()
    result["msku"] = result["msku"].str.replace(r"^\t+", "", regex=True)
    result["人员"] = result["开发专员"].map(normalize_department_person_name)
    result["店铺前缀"] = result["店铺"].map(extract_department_store_prefix)
    result["店铺编码"] = result["店铺"].map(extract_department_store_code)
    result["店铺部门"] = result["店铺"].map(department_name_from_store)
    for col in date_columns:
        result[col] = normalize_config_number(result[col]).fillna(0)
    return result


def normalize_department_person_name(value) -> str:
    if pd.isna(value):
        return ""
    text = normalize_department_developer_text(value)
    if not text or text == "--":
        return ""
    text = re.sub(r"^运营[一二三四五六七八九十百千万0-9]+部-", "", text)
    text = re.sub(r"-?26$", "", text)
    return text.strip()


def normalize_department_developer_text(value) -> str:
    text = unicodedata.normalize("NFKC", str(value)).strip()
    text = re.sub(r"[\u2010-\u2015\u2212－]", "-", text)
    text = re.sub(r"\s*-\s*", "-", text)
    return text


def extract_department_store_prefix(value) -> str:
    text = normalize_department_developer_text(value)
    match = re.match(r"^\s*(\d+)\s*-", text)
    return match.group(1) if match else ""


def extract_department_store_code(value) -> str:
    """Return the normalized store code for a ``20-`` store, otherwise blank."""
    text = normalize_department_developer_text(value)
    match = re.match(r"^\s*20\s*-\s*([A-Za-z0-9]+)", text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def department_name_from_store(value) -> str:
    prefix = extract_department_store_prefix(value)
    if prefix in {"6", "7"}:
        return "联合部门"
    if prefix == "20":
        return "运营二十部"
    return ""


def department_metric_columns_for_dates(dates: list[pd.Timestamp], suffix: str) -> list[str]:
    return [f"{date.strftime('%m-%d')}{suffix}" for date in dates]


def latest_department_detail_date(
    volume_df: pd.DataFrame,
    amount_df: pd.DataFrame,
    reference_date=None,
) -> pd.Timestamp | None:
    """Return the latest date shared by the department volume and amount sources."""
    volume_days = {
        match.group(1)
        for column in volume_df.columns
        if (match := re.fullmatch(r"(\d{2}-\d{2})销量", str(column)))
    }
    amount_days = {
        match.group(1)
        for column in amount_df.columns
        if (match := re.fullmatch(r"(\d{2}-\d{2})销售额", str(column)))
    }
    common_days = volume_days & amount_days
    if not common_days:
        return None

    reference = pd.Timestamp(reference_date).normalize() if reference_date is not None else pd.Timestamp.today().normalize()
    candidates = []
    for day in common_days:
        month, day_of_month = map(int, day.split("-"))
        try:
            candidate = pd.Timestamp(year=reference.year, month=month, day=day_of_month)
            if candidate > reference:
                candidate = candidate.replace(year=reference.year - 1)
            candidates.append(candidate)
        except ValueError:
            continue
    return max(candidates) if candidates else None


def department_month_dates(today: pd.Timestamp) -> list[pd.Timestamp]:
    month_start = today.replace(day=1)
    if today.day == 1:
        return []
    return list(pd.date_range(month_start, today - pd.Timedelta(days=1), freq="D"))


def department_performance_daily_columns(dates: list[pd.Timestamp]) -> list[str]:
    columns = []
    for date in dates:
        label = f"{date.month}月{date.day}日"
        columns.extend([f"{label}销量", f"{label}销售额（元）"])
    return columns


def department_performance_columns(label_col: str, dates: list[pd.Timestamp]) -> list[str]:
    return [label_col] + DEPARTMENT_PERFORMANCE_FIXED_COLUMNS + department_performance_daily_columns(dates)


def with_department_performance_total(frame: pd.DataFrame) -> pd.DataFrame:
    """Prepend a total row for either department or developer performance."""
    if frame.empty:
        return frame
    label_col = next((column for column in ("部门", "店铺", "开发员") if column in frame.columns), "开发员")
    total = {label_col: "合计"}
    for column in frame.columns:
        if column != label_col:
            total[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0).sum()
    return pd.concat([pd.DataFrame([total]), frame], ignore_index=True)


def build_department_performance_tables(
    operational_df: pd.DataFrame,
    volume_df: pd.DataFrame,
    amount_df: pd.DataFrame,
    today=None,
) -> dict[str, pd.DataFrame]:
    # Official exports can contain byte-for-byte repeated detail lines. Remove
    # only rows whose normalized calculation fields are completely identical;
    # rows for the same SKU with any differing value remain independent.
    volume = normalize_sales_volume_detail(volume_df).drop_duplicates(keep="first").reset_index(drop=True)
    amount = normalize_sales_amount_detail(amount_df).drop_duplicates(keep="first").reset_index(drop=True)
    today_ts = pd.Timestamp(today).normalize() if today is not None else latest_department_detail_date(volume, amount)
    if today_ts is None:
        raise ValueError("销量明细与销售额明细没有共同的日期列")
    window_dates = [today_ts - pd.Timedelta(days=offset) for offset in range(1, 8)]
    volume_date_cols = department_metric_columns_for_dates(window_dates, "销量")
    amount_date_cols = department_metric_columns_for_dates(window_dates, "销售额")
    missing_volume = [col for col in volume_date_cols if col not in volume.columns]
    missing_amount = [col for col in amount_date_cols if col not in amount.columns]
    if missing_volume:
        raise ValueError(f"销量明细缺少近7天日期列：{', '.join(missing_volume)}")
    if missing_amount:
        raise ValueError(f"销售额明细缺少近7天日期列：{', '.join(missing_amount)}")

    month_dates = department_month_dates(today_ts)
    month_amount_cols = [col for col in department_metric_columns_for_dates(month_dates, "销售额") if col in amount.columns]
    remaining_days = calendar.monthrange(today_ts.year, today_ts.month)[1] - today_ts.day + 1
    onsale_counts = build_department_onsale_counts(operational_df)
    department_volume = volume[volume["店铺部门"].ne("")].copy()
    department_amount = amount[amount["店铺部门"].ne("")].copy()
    store_codes = build_department_store_codes(operational_df)
    store_volume = volume[volume["店铺编码"].isin(store_codes)].copy()
    store_amount = amount[amount["店铺编码"].isin(store_codes)].copy()
    return {
        "开发员业绩排行": build_department_performance_table_for_group(
            "开发员",
            "人员",
            "person",
            volume,
            amount,
            window_dates,
            volume_date_cols,
            amount_date_cols,
            month_amount_cols,
            remaining_days,
            onsale_counts,
        ),
        "部门业绩": build_department_performance_table_for_group(
            "部门",
            "店铺部门",
            "department",
            department_volume,
            department_amount,
            window_dates,
            volume_date_cols,
            amount_date_cols,
            month_amount_cols,
            remaining_days,
            onsale_counts,
        ),
        "店铺业绩排行": build_department_performance_table_for_group(
            "店铺",
            "店铺编码",
            "store",
            store_volume,
            store_amount,
            window_dates,
            volume_date_cols,
            amount_date_cols,
            month_amount_cols,
            remaining_days,
            onsale_counts,
            labels=store_codes,
        ),
    }


def build_department_performance_table_for_group(
    label_col: str,
    group_col: str,
    count_kind: str,
    volume: pd.DataFrame,
    amount: pd.DataFrame,
    window_dates: list[pd.Timestamp],
    volume_date_cols: list[str],
    amount_date_cols: list[str],
    month_amount_cols: list[str],
    remaining_days: int,
    onsale_counts: dict[tuple[str, str | None], int],
    labels: Iterable[str] | None = None,
) -> pd.DataFrame:
    if labels is None:
        labels = {label for label in volume[group_col].tolist() + amount[group_col].tolist() if label}
    else:
        labels = {label for label in labels if label}
    labels = sorted(labels)
    rows = []
    for label in labels:
        rows.append(
            build_department_performance_row(
                label_col,
                label,
                count_kind,
                volume[volume[group_col].eq(label)].copy(),
                amount[amount[group_col].eq(label)].copy(),
                window_dates,
                volume_date_cols,
                amount_date_cols,
                month_amount_cols,
                remaining_days,
                onsale_counts,
            )
        )
    if not rows:
        return pd.DataFrame(columns=department_performance_columns(label_col, window_dates))
    rows = sorted(rows, key=lambda row: row["近7天日均销售额（元）"], reverse=True)
    result = pd.DataFrame(rows)
    denominator = result["近7天日均销售额（元）"].sum()
    result["销售额贡献占比"] = result["近7天日均销售额（元）"].map(lambda value: safe_blank_ratio(value, denominator))
    return result[department_performance_columns(label_col, window_dates)].reset_index(drop=True)


def build_department_performance_row(
    label_col: str,
    label: str,
    count_kind: str,
    volume: pd.DataFrame,
    amount: pd.DataFrame,
    window_dates: list[pd.Timestamp],
    volume_date_cols: list[str],
    amount_date_cols: list[str],
    month_amount_cols: list[str],
    remaining_days: int,
    onsale_counts: dict[tuple[str, str | None], int],
) -> dict:
    total_volume_7 = volume[volume_date_cols].sum().sum() if not volume.empty else 0
    total_amount_7 = amount[amount_date_cols].sum().sum() if not amount.empty else 0
    month_amount = amount[month_amount_cols].sum().sum() if month_amount_cols and not amount.empty else 0
    row = {
        label_col: label,
        "在售SKU数量": onsale_counts.get((count_kind, label), 0),
        "销售额贡献占比": 0,
        "近7天日均订单": total_volume_7 / 7,
        "近7天日均销售额（元）": total_amount_7 / 7,
        "预估本月销售额（元）": month_amount + (total_amount_7 / 7) * remaining_days,
    }
    for date, volume_col, amount_col in zip(window_dates, volume_date_cols, amount_date_cols):
        label_prefix = f"{date.month}月{date.day}日"
        row[f"{label_prefix}销量"] = volume[volume_col].sum() if not volume.empty else 0
        row[f"{label_prefix}销售额（元）"] = amount[amount_col].sum() if not amount.empty else 0
    return row


def build_department_onsale_counts(operational_df: pd.DataFrame) -> dict[tuple[str, str | None], int]:
    required = ["MSKU", "店铺名称", "开发员", "可售"]
    missing = [col for col in required if col not in operational_df.columns]
    if missing:
        raise ValueError(f"运营原始表缺少部门监控在售列：{', '.join(missing)}")

    counts: dict[tuple[str, str | None], set[str]] = {}
    base = operational_df[required].copy()
    base["MSKU"] = base["MSKU"].fillna("").astype(str).str.strip().str.upper()
    base["开发员"] = base["开发员"].map(normalize_department_person_name)
    base["店铺部门"] = base["店铺名称"].map(department_name_from_store)
    base["可售"] = normalize_config_number(base["可售"]).fillna(0)
    base = base[base["MSKU"].ne("") & base["可售"].gt(0)].copy()
    for _, row in base.iterrows():
        if row["开发员"]:
            counts.setdefault(("person", row["开发员"]), set()).add(row["MSKU"])
        if row["店铺部门"]:
            counts.setdefault(("department", row["店铺部门"]), set()).add(row["MSKU"])
        for _, store_name in extract_operational_store_codes(row["店铺名称"]):
            store_code = extract_department_store_code(store_name)
            if store_code:
                counts.setdefault(("store", store_code), set()).add(row["MSKU"])
    return {key: len(value) for key, value in counts.items()}


def build_department_store_codes(operational_df: pd.DataFrame) -> list[str]:
    """Return all unique ``20-`` store codes present in the operational source."""
    if "店铺名称" not in operational_df.columns:
        raise ValueError("运营原始表缺少部门监控店铺列：店铺名称")
    codes: set[str] = set()
    for value in operational_df["店铺名称"].fillna("").astype(str):
        for _, store_name in extract_operational_store_codes(value):
            code = extract_department_store_code(store_name)
            if code:
                codes.add(code)
    return sorted(codes)


def normalize_replenishment_targets(targets: pd.DataFrame | None) -> pd.DataFrame:
    if targets is None or targets.empty:
        return pd.DataFrame(columns=REPLENISHMENT_TARGET_COLUMNS)

    data = targets.copy()
    for col in REPLENISHMENT_TARGET_COLUMNS:
        if col not in data.columns:
            data[col] = pd.NA
    optional_columns = [column for column in ["箱规"] if column in data.columns]
    data = data[REPLENISHMENT_TARGET_COLUMNS + optional_columns].copy()
    data["ASIN"] = data["ASIN"].fillna("").astype(str).str.strip()
    data["目标可售天数"] = normalize_config_number(data["目标可售天数"]).round()
    if "箱规" in data.columns:
        data["箱规"] = normalize_config_number(data["箱规"]).round()
        data["箱规"] = data["箱规"].where(data["箱规"].gt(0))
    configured = data["目标可售天数"].notna()
    if "箱规" in data.columns:
        configured |= data["箱规"].notna()
    data = data[data["ASIN"].ne("") & configured].copy()
    data["目标可售天数"] = data["目标可售天数"].clip(lower=0).astype("Int64")
    return data.drop_duplicates(subset=["ASIN"], keep="last").reset_index(drop=True)


def normalize_replenishment_operational(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in REPLENISHMENT_OPERATIONAL_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"运营原始表缺少补货管理列：{', '.join(missing)}")

    optional_columns = AGING_STOCK_COLUMNS + ["备注"]
    base = df[
        REPLENISHMENT_OPERATIONAL_REQUIRED_COLUMNS
        + [column for column in optional_columns if column in df.columns]
    ].copy()
    for column in optional_columns:
        if column not in base.columns:
            base[column] = pd.NA
    for col in ["ASIN", "MSKU", "店铺名称", "开发员"]:
        base[col] = base[col].fillna("").astype(str).str.strip()
    base["ASIN"] = base["ASIN"].str.upper()
    base["备注"] = base["备注"].fillna("").astype(str).str.strip()
    for col in REPLENISHMENT_SALES_COLUMNS + REPLENISHMENT_FORMULA_STOCK_COLUMNS + AGING_STOCK_COLUMNS + ["单品重量(g)"]:
        # Do not replace invalid/missing inputs with zero.  A replenishment
        # recommendation is unsafe unless every calculation input is present.
        base[col] = normalize_config_number(base[col])
    base["上架时间"] = normalize_replenishment_listing_dates(base["上架时间"])
    return base


def normalize_replenishment_listing_dates(values: pd.Series) -> pd.Series:
    """Parse listing dates as local calendar dates without timezone shifting.

    Operational exports currently mix Excel datetimes, European
    ``DD/MM/YYYY`` strings and ISO ``YYYY-MM-DD`` strings, commonly followed
    by MET/MEST/CET/CEST/GMT.  Replenishment age is calendar-day based, so the
    suffix must not move a record into another date through UTC conversion.
    """

    result = pd.Series(pd.NaT, index=values.index, dtype="datetime64[ns]")
    non_empty = values.notna() & values.astype(str).str.strip().ne("")
    if not non_empty.any():
        return result

    numeric = pd.to_numeric(values.where(non_empty), errors="coerce")
    excel_serial = numeric.between(20_000, 80_000, inclusive="both")
    if excel_serial.any():
        result.loc[excel_serial] = pd.to_datetime(
            numeric.loc[excel_serial],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        ).dt.normalize()

    text = values.where(non_empty & ~excel_serial).astype("string").str.strip()
    text = text.str.replace(r"\s+(?:MET|MEST|CET|CEST|GMT)$", "", regex=True, case=False)
    iso_date = text.str.extract(r"^(\d{4}-\d{2}-\d{2})", expand=False)
    european_date = text.str.extract(r"^(\d{2}/\d{2}/\d{4})", expand=False)
    iso_mask = iso_date.notna()
    european_mask = european_date.notna() & ~iso_mask
    if iso_mask.any():
        result.loc[iso_mask] = pd.to_datetime(
            iso_date.loc[iso_mask], format="%Y-%m-%d", errors="coerce"
        )
    if european_mask.any():
        result.loc[european_mask] = pd.to_datetime(
            european_date.loc[european_mask], format="%d/%m/%Y", errors="coerce"
        )

    fallback_mask = non_empty & result.isna() & ~excel_serial
    if fallback_mask.any():
        fallback = pd.to_datetime(
            text.loc[fallback_mask],
            errors="coerce",
            format="mixed",
            dayfirst=True,
        )
        if isinstance(fallback.dtype, pd.DatetimeTZDtype):
            fallback = fallback.dt.tz_localize(None)
        result.loc[fallback_mask] = fallback
    return result.dt.normalize()


def build_replenishment_operational_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Legacy flat summary retained for callers outside the new page.

    The replenishment page itself uses ``build_replenishment_management_tables``
    so it can retain SKU evidence and data-quality status.
    """
    operational = normalize_replenishment_operational(df)
    if operational.empty:
        return pd.DataFrame(columns=replenishment_operational_columns())

    grouped = operational.groupby("ASIN", dropna=False, sort=False).agg(
        MSKU=("MSKU", join_non_empty_values),
        店铺编码=("店铺名称", join_operational_store_codes),
        开发员=("开发员", join_non_empty_values),
        重量=("单品重量(g)", "max"),
    ).reset_index()
    numeric = operational.groupby("ASIN", dropna=False, sort=False)[REPLENISHMENT_FORMULA_STOCK_COLUMNS].sum(min_count=1)
    grouped["亚马逊可售库存数量"] = numeric[["可售", "待调仓", "调仓中", "待入库"]].sum(axis=1).to_numpy()
    grouped["总库存数量"] = (
        numeric["可售"] + numeric["待调仓"] + numeric["调仓中"] + numeric["待入库"] * 2
        + numeric["采购在途"] + numeric["本地库存"] + numeric["在途"] + numeric["计划入库"]
    ).to_numpy()
    grouped["日均销量"] = (
        operational.groupby("ASIN", dropna=False, sort=False)[REPLENISHMENT_SALES_COLUMNS].sum(min_count=1).sum(axis=1) / 30
    ).to_numpy()
    grouped["库龄超90天库存数"] = pd.NA
    grouped["建议补货方式"] = grouped["重量"].map(lambda value: "空运" if pd.notna(value) and value < 100 else "卡航")
    return grouped[replenishment_operational_columns()].reset_index(drop=True)


def join_operational_store_codes(values: pd.Series) -> str:
    codes = []
    seen = set()
    for value in values:
        for code, _ in extract_operational_store_codes(value):
            if code and code not in seen:
                codes.append(code)
                seen.add(code)
    return "；".join(codes)


def replenishment_operational_columns() -> list[str]:
    return [
        "ASIN",
        "MSKU",
        "店铺编码",
        "开发员",
        "亚马逊可售库存数量",
        "总库存数量",
        "库龄超90天库存数",
        "日均销量",
        "重量",
        "建议补货方式",
    ]


def default_replenishment_coverage_rules() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"运输方式": "空运", "重量下限": 0, "重量上限": 100, "头程时效": 30, "预警天数": 40, "补货频次": 10, "是否启用": True},
            {"运输方式": "卡航", "重量下限": 100, "重量上限": pd.NA, "头程时效": 40, "预警天数": 40, "补货频次": 10, "是否启用": True},
        ],
        columns=REPLENISHMENT_COVERAGE_RULE_COLUMNS,
    )


def normalize_replenishment_coverage_rules(rules: pd.DataFrame | None) -> pd.DataFrame:
    if rules is None or rules.empty:
        return pd.DataFrame(columns=REPLENISHMENT_COVERAGE_RULE_COLUMNS)
    data = rules.copy()
    for column in REPLENISHMENT_COVERAGE_RULE_COLUMNS:
        if column not in data.columns:
            data[column] = pd.NA
    data = data[REPLENISHMENT_COVERAGE_RULE_COLUMNS].copy()
    data["运输方式"] = data["运输方式"].fillna("").astype(str).str.strip()
    for column in ["重量下限", "重量上限", "头程时效", "预警天数", "补货频次"]:
        data[column] = normalize_config_number(data[column])
    data["是否启用"] = data["是否启用"].map(is_enabled)
    invalid = data["运输方式"].eq("") | data["重量下限"].isna() | data[["头程时效", "预警天数", "补货频次"]].isna().any(axis=1)
    if invalid.any():
        raise ValueError("库存覆盖规则必须填写运输方式、重量下限、头程时效、预警天数和补货频次")
    if (data[["重量下限", "头程时效", "预警天数", "补货频次"]] < 0).any().any():
        raise ValueError("库存覆盖规则的重量和天数不能为负数")
    bounded = data["重量上限"].notna()
    if (data.loc[bounded, "重量上限"] < data.loc[bounded, "重量下限"]).any():
        raise ValueError("库存覆盖规则的重量上限不能小于重量下限")
    enabled = data[data["是否启用"]].sort_values(["重量下限", "重量上限"], na_position="last", kind="stable")
    if enabled.empty:
        raise ValueError("至少需要启用一条库存覆盖规则")
    if float(enabled.iloc[0]["重量下限"]) != 0:
        raise ValueError("启用的库存覆盖规则必须从0g开始")
    last_upper: float | None = None
    for _, row in enabled.iterrows():
        lower = float(row["重量下限"])
        if last_upper is not None and lower < last_upper:
            raise ValueError("启用的库存覆盖规则重量区间不能重叠")
        if last_upper is not None and lower > last_upper:
            raise ValueError("启用的库存覆盖规则重量区间不能留空档")
        upper = row["重量上限"]
        last_upper = None if pd.isna(upper) else float(upper)
    if last_upper is not None:
        raise ValueError("启用的库存覆盖规则最后一条必须不填写重量上限")
    return data.reset_index(drop=True)


def normalize_replenishment_switches(switches: pd.DataFrame | None) -> pd.DataFrame:
    if switches is None or switches.empty:
        return pd.DataFrame(columns=REPLENISHMENT_SWITCH_COLUMNS)
    data = switches.copy()
    if "ASIN" not in data.columns and "补货组ID" in data.columns:
        data = data.rename(columns={"补货组ID": "ASIN"})
    elif "ASIN" in data.columns and "补货组ID" in data.columns:
        asin = data["ASIN"].fillna("").astype(str).str.strip()
        data.loc[asin.eq(""), "ASIN"] = data.loc[asin.eq(""), "补货组ID"]
    for column in REPLENISHMENT_SWITCH_COLUMNS:
        if column not in data.columns:
            data[column] = pd.NA
    data = data[REPLENISHMENT_SWITCH_COLUMNS].copy()
    data["ASIN"] = data["ASIN"].fillna("").astype(str).str.strip().str.upper()
    data["关闭原因"] = data["关闭原因"].fillna("").astype(str).str.strip()
    data["是否补货"] = data["是否补货"].map(is_enabled)
    if data["ASIN"].eq("").any():
        raise ValueError("补货开关必须填写ASIN")
    if ((~data["是否补货"]) & data["关闭原因"].eq("")).any():
        raise ValueError("关闭补货时必须填写关闭原因")
    return data.drop_duplicates(subset=["ASIN"], keep="last").reset_index(drop=True)


def normalize_replenishment_product_tags(tags: pd.DataFrame | None) -> pd.DataFrame:
    if tags is None or tags.empty:
        return pd.DataFrame(columns=REPLENISHMENT_PRODUCT_TAG_COLUMNS)
    data = tags.copy()
    for column in REPLENISHMENT_PRODUCT_TAG_COLUMNS:
        if column not in data.columns:
            data[column] = pd.NA
    data = data[REPLENISHMENT_PRODUCT_TAG_COLUMNS].copy()
    for column in ["ASIN", "产品标签", "标签颜色", "备注"]:
        data[column] = data[column].fillna("").astype(str).str.strip()
    data["ASIN"] = data["ASIN"].str.upper()
    data["是否启用"] = data["是否启用"].map(is_enabled)
    if data["ASIN"].eq("").any() or data["产品标签"].eq("").any():
        raise ValueError("ASIN产品标签必须填写ASIN和产品标签")
    invalid_colors = data["标签颜色"].ne("") & ~data["标签颜色"].str.fullmatch(r"#[0-9A-Fa-f]{6}")
    if invalid_colors.any():
        raise ValueError("标签颜色必须为空或使用#RRGGBB格式")
    return data.drop_duplicates(subset=["ASIN", "产品标签"], keep="last").reset_index(drop=True)


def normalize_sales_history_2025(df: pd.DataFrame) -> pd.DataFrame:
    """Validate the canonical ASIN-level summary derived from the 12-sheet workbook."""
    missing = [column for column in SALES_HISTORY_2025_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"25年销量明细汇总缺少列：{', '.join(missing)}")
    data = df[SALES_HISTORY_2025_COLUMNS].copy()
    data["ASIN"] = data["ASIN"].fillna("").astype(str).str.strip().str.upper()
    if data["ASIN"].eq("").any():
        raise ValueError("25年销量明细汇总存在空ASIN")
    if data["ASIN"].duplicated().any():
        raise ValueError("25年销量明细汇总的ASIN必须唯一")
    for column in SALES_HISTORY_2025_COLUMNS[1:]:
        original = data[column]
        numeric = pd.to_numeric(original, errors="coerce")
        invalid = original.notna() & original.astype(str).str.strip().ne("") & numeric.isna()
        if invalid.any():
            raise ValueError(f"25年销量明细汇总列“{column}”存在非数字")
        data[column] = numeric.fillna(0)
    return data.reset_index(drop=True)


def normalize_sales_history_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Validate the ASIN-level, fixed-width history used by replenishment."""
    missing = [column for column in SALES_HISTORY_GENERIC_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"销量历史汇总缺少列：{', '.join(missing)}")
    data = df[SALES_HISTORY_GENERIC_COLUMNS].copy()
    data["ASIN"] = data["ASIN"].fillna("").astype(str).str.strip().str.upper()
    if data["ASIN"].eq("").any():
        raise ValueError("销量历史汇总存在空ASIN")
    if data["ASIN"].duplicated().any():
        raise ValueError("销量历史汇总的ASIN必须唯一")
    for column in [f"历史月份{index}" for index in range(1, 13)]:
        values = data[column].fillna("").astype(str).str.strip()
        invalid = values.ne("") & ~values.str.fullmatch(r"\d{4}-(0[1-9]|1[0-2])")
        if invalid.any():
            raise ValueError(f"销量历史汇总列“{column}”存在非法月份")
        data[column] = values
    for column in [column for column in SALES_HISTORY_GENERIC_COLUMNS if column != "ASIN" and not column.startswith("历史月份")]:
        original = data[column]
        numeric = pd.to_numeric(original, errors="coerce")
        invalid = original.notna() & original.astype(str).str.strip().ne("") & numeric.isna()
        if invalid.any():
            raise ValueError(f"销量历史汇总列“{column}”存在非数字")
        data[column] = numeric.fillna(0)
    for index in range(1, 13):
        data[f"历史{index}月计入天数"] = data[f"历史{index}月计入天数"].round().astype(int)
    return data.reset_index(drop=True)


def normalize_sales_history_month_source(df: pd.DataFrame, month: str) -> pd.DataFrame:
    """Validate one complete monthly sales-history CSV without changing the raw upload."""

    match = re.fullmatch(r"(\d{4})-(0[1-9]|1[0-2])", str(month).strip())
    if not match:
        raise ValueError(f"销量历史月份必须是合法的 YYYY-MM：{month}")
    year, month_number = int(match.group(1)), int(match.group(2))
    data = df.copy()
    data.columns = [str(column).strip().lstrip("\ufeff") for column in data.columns]
    if data.empty:
        raise ValueError(f"{month}销量历史没有数据行")
    duplicate_columns = data.columns[data.columns.duplicated()].astype(str).tolist()
    if duplicate_columns:
        raise ValueError(f"{month}销量历史包含重复列名：{', '.join(duplicate_columns[:10])}")
    required = ["asin", "msku", "国家", "小计"]
    day_columns = [f"{month_number:02d}-{day:02d}销量" for day in range(1, calendar.monthrange(year, month_number)[1] + 1)]
    missing = [column for column in [*required, *day_columns] if column not in data.columns]
    if missing:
        raise ValueError(f"{month}销量历史缺少列：{', '.join(missing[:12])}")
    actual_day_columns = {
        str(column) for column in data.columns if re.fullmatch(r"\d{2}-\d{2}销量", str(column))
    }
    unexpected = sorted(actual_day_columns.difference(day_columns))
    if unexpected:
        raise ValueError(f"{month}销量历史包含不属于该月的销量列：{', '.join(unexpected[:10])}")

    result = data.copy()
    result["asin"] = result["asin"].fillna("").astype(str).str.strip().str.upper()
    if result["asin"].eq("").any():
        row_number = int(result.index[result["asin"].eq("")][0]) + 2
        raise ValueError(f"{month}销量历史第{row_number}行ASIN为空")
    numeric_columns = ["小计", *day_columns]
    for column in numeric_columns:
        original = result[column]
        values = pd.to_numeric(original, errors="coerce")
        invalid = original.notna() & original.astype(str).str.strip().ne("") & values.isna()
        if invalid.any():
            row_number = int(result.index[invalid][0]) + 2
            raise ValueError(f"{month}销量历史第{row_number}行“{column}”不是有效数字")
        if values.fillna(0).lt(0).any():
            row_number = int(result.index[values.fillna(0).lt(0)][0]) + 2
            raise ValueError(f"{month}销量历史第{row_number}行“{column}”不能为负数")
        result[column] = values.fillna(0).astype(float)
    daily_total = result[day_columns].sum(axis=1)
    mismatch = ~daily_total.sub(result["小计"]).abs().le(1e-6)
    if mismatch.any():
        row_number = int(result.index[mismatch][0]) + 2
        raise ValueError(f"{month}销量历史第{row_number}行小计与每日销量合计不一致")
    return result


def _history_number(value, *, sheet_name: str, row_number: int, column: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        return 0.0
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{sheet_name}第{row_number}行“{column}”不是有效数字：{value}") from exc
    if not math.isfinite(number):
        raise ValueError(f"{sheet_name}第{row_number}行“{column}”不是有限数字")
    if number < 0:
        raise ValueError(f"{sheet_name}第{row_number}行“{column}”不能为负数")
    return number


def _history_output_number(value: float) -> int | float:
    return int(round(value)) if math.isclose(value, round(value), abs_tol=1e-9) else round(value, 6)


def _month_label(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _shift_month(year: int, month: int, offset: int) -> tuple[int, int]:
    month_index = year * 12 + month - 1 + offset
    return month_index // 12, month_index % 12 + 1


def _rolling_window_labels(today: date | datetime | pd.Timestamp | None = None) -> list[str]:
    current = pd.Timestamp(today).date() if today is not None else pd.Timestamp.now(tz="Asia/Shanghai").date()
    anchor_year, anchor_month = _shift_month(current.year, current.month, -1)
    return [
        _month_label(*_shift_month(anchor_year, anchor_month, offset))
        for offset in range(-11, 1)
    ]


def _build_sales_history_workbook_summary(
    source: Path | bytes | bytearray | io.BytesIO,
    sheet_dates: list[tuple[str, int, int]],
    *,
    title: str,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Parse a twelve-sheet workbook and apply the cross-month stockout rule."""
    from openpyxl import load_workbook

    workbook_source = io.BytesIO(bytes(source)) if isinstance(source, (bytes, bytearray)) else source
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        actual_sheets = set(workbook.sheetnames)
        expected_sheets = {name for name, _, _ in sheet_dates}
        missing_sheets = sorted(expected_sheets.difference(actual_sheets))
        unexpected_sheets = sorted(actual_sheets.difference(expected_sheets))
        if missing_sheets or unexpected_sheets or len(actual_sheets) != 12:
            issues = []
            if missing_sheets:
                issues.append(f"缺少sheet：{', '.join(missing_sheets)}")
            if unexpected_sheets:
                issues.append(f"存在额外sheet：{', '.join(unexpected_sheets)}")
            raise ValueError(f"{title}必须且只能包含连续12个月的sheet；{'；'.join(issues)}")

        all_asins: set[str] = set()
        site_totals: dict[str, dict[str, float]] = {}
        daily_by_asin: dict[str, dict[date, float]] = {}
        total_rows = 0
        maximum_columns = 0

        for sheet_name, year, month in sheet_dates:
            sheet = workbook[sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [str(value).strip() if value is not None else "" for value in header_row]
            maximum_columns = max(maximum_columns, len(headers))
            duplicate_headers = sorted({name for name in headers if name and headers.count(name) > 1})
            if duplicate_headers:
                raise ValueError(f"{sheet_name}存在重复列名：{', '.join(duplicate_headers[:10])}")
            header_index = {name: index for index, name in enumerate(headers)}
            expected_days = calendar.monthrange(year, month)[1]
            day_columns = [f"{month:02d}-{day:02d}销量" for day in range(1, expected_days + 1)]
            required_columns = ["asin", "msku", "国家", "小计", *day_columns]
            missing_columns = [column for column in required_columns if column not in header_index]
            if missing_columns:
                raise ValueError(f"{sheet_name}缺少列：{', '.join(missing_columns[:12])}")
            actual_day_columns = {
                name for name in headers if re.fullmatch(r"\d{2}-\d{2}销量", name)
            }
            unexpected_days = sorted(actual_day_columns.difference(day_columns))
            if unexpected_days:
                raise ValueError(f"{sheet_name}包含不属于该月的销量列：{', '.join(unexpected_days[:10])}")

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                total_rows += 1
                asin = str(row[header_index["asin"]] or "").strip().upper()
                if not asin:
                    raise ValueError(f"{sheet_name}第{row_number}行ASIN为空")
                all_asins.add(asin)
                daily_values = [
                    _history_number(
                        row[header_index[column]],
                        sheet_name=sheet_name,
                        row_number=row_number,
                        column=column,
                    )
                    for column in day_columns
                ]
                subtotal = _history_number(
                    row[header_index["小计"]],
                    sheet_name=sheet_name,
                    row_number=row_number,
                    column="小计",
                )
                daily_sum = sum(daily_values)
                if not math.isclose(subtotal, daily_sum, abs_tol=1e-6):
                    raise ValueError(
                        f"{sheet_name}第{row_number}行小计与每日销量合计不一致："
                        f"小计={_history_output_number(subtotal)}，每日合计={_history_output_number(daily_sum)}"
                    )

                country = str(row[header_index["国家"]] or "").strip()
                country_code = SALES_HISTORY_2025_COUNTRIES.get(country)
                if country_code is None:
                    continue
                asin_sites = site_totals.setdefault(
                    asin, {code: 0.0 for code in SALES_HISTORY_2025_COUNTRIES.values()}
                )
                asin_sites[country_code] += daily_sum
                asin_daily = daily_by_asin.setdefault(asin, {})
                for index, value in enumerate(daily_values):
                    asin_daily[date(year, month, index + 1)] = asin_daily.get(date(year, month, index + 1), 0.0) + value

        all_dates = [
            date(year, month, day)
            for _, year, month in sheet_dates
            for day in range(1, calendar.monthrange(year, month)[1] + 1)
        ]
        month_labels = [_month_label(year, month) for _, year, month in sheet_dates]
        month_dates = [
            [date(year, month, day) for day in range(1, calendar.monthrange(year, month)[1] + 1)]
            for _, year, month in sheet_dates
        ]

        def excluded_stockout_dates(values: list[float]) -> set[date]:
            excluded: set[date] = set()
            run_start: int | None = None
            for index, value in enumerate(values + [1.0]):
                if value == 0 and run_start is None:
                    run_start = index
                elif value != 0 and run_start is not None:
                    if index - run_start >= 10:
                        excluded.update(all_dates[run_start:index])
                    run_start = None
            return excluded

        rows: list[dict[str, object]] = []
        for asin in sorted(all_asins):
            row: dict[str, object] = {"ASIN": asin}
            asin_sites = site_totals.get(
                asin, {code: 0.0 for code in SALES_HISTORY_2025_COUNTRIES.values()}
            )
            for code in SALES_HISTORY_2025_COUNTRIES.values():
                row[f"{code}总销量"] = _history_output_number(asin_sites[code])
            daily_values = [daily_by_asin.get(asin, {}).get(day, 0.0) for day in all_dates]
            excluded = excluded_stockout_dates(daily_values)
            for index, (label, dates) in enumerate(zip(month_labels, month_dates), start=1):
                total_sales = sum(daily_by_asin.get(asin, {}).get(day, 0.0) for day in dates)
                included_days = sum(day not in excluded for day in dates)
                adjusted_average = round(total_sales / included_days, 2) if total_sales and included_days else 0.0
                row[f"历史月份{index}"] = label
                row[f"历史{index}月总销量"] = _history_output_number(total_sales)
                row[f"历史{index}月计入天数"] = included_days
                row[f"历史{index}月日均销量"] = adjusted_average
            rows.append(row)
        summary = normalize_sales_history_summary(pd.DataFrame(rows, columns=SALES_HISTORY_GENERIC_COLUMNS))
        return summary, {
            "rows": total_rows,
            "effective_rows": len(summary),
            "columns": maximum_columns,
        }
    finally:
        workbook.close()


def _generic_to_legacy_history(summary: pd.DataFrame) -> pd.DataFrame:
    data = normalize_sales_history_summary(summary)
    rows: list[dict[str, object]] = []
    for _, source_row in data.iterrows():
        row: dict[str, object] = {"ASIN": source_row["ASIN"]}
        for code in SALES_HISTORY_2025_COUNTRIES.values():
            row[f"{code}总销量"] = source_row[f"{code}总销量"]
        for index in range(1, 13):
            row[f"{index}月总销量"] = source_row[f"历史{index}月总销量"]
            row[f"{index}月出单天数"] = source_row[f"历史{index}月计入天数"]
            row[f"{index}月除0日均"] = source_row[f"历史{index}月日均销量"]
        rows.append(row)
    return normalize_sales_history_2025(pd.DataFrame(rows, columns=SALES_HISTORY_2025_COLUMNS))


def _legacy_to_generic_history(summary: pd.DataFrame) -> pd.DataFrame:
    data = normalize_sales_history_2025(summary)
    rows: list[dict[str, object]] = []
    for _, source_row in data.iterrows():
        row: dict[str, object] = {"ASIN": source_row["ASIN"]}
        for code in SALES_HISTORY_2025_COUNTRIES.values():
            row[f"{code}总销量"] = source_row[f"{code}总销量"]
        for index in range(1, 13):
            row[f"历史月份{index}"] = f"2025-{index:02d}"
            row[f"历史{index}月总销量"] = source_row[f"{index}月总销量"]
            row[f"历史{index}月计入天数"] = source_row[f"{index}月出单天数"]
            row[f"历史{index}月日均销量"] = source_row[f"{index}月除0日均"]
        rows.append(row)
    return normalize_sales_history_summary(pd.DataFrame(rows, columns=SALES_HISTORY_GENERIC_COLUMNS))


def _history_with_legacy_aliases(summary: pd.DataFrame) -> pd.DataFrame:
    """Keep the old monthly columns in the supporting return frame only."""
    data = normalize_sales_history_summary(summary).copy()
    legacy = _generic_to_legacy_history(data)
    for column in SALES_HISTORY_2025_MONTH_COLUMNS:
        data[column] = legacy[column]
    return data


def build_sales_history_2025_summary(
    source: Path | bytes | bytearray | io.BytesIO,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Parse the legacy 1–12月 workbook using the corrected stockout rule."""
    sheet_dates = [(f"{month}月", 2025, month) for month in range(1, 13)]
    generic, stats = _build_sales_history_workbook_summary(source, sheet_dates, title="25年销量明细")
    return _generic_to_legacy_history(generic), stats


def build_sales_history_rolling_summary(
    source: Path | bytes | bytearray | io.BytesIO,
    *,
    today: date | datetime | pd.Timestamp | None = None,
    require_latest: bool = False,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Parse a rolling YYYY-MM workbook and return the generic history summary."""
    from openpyxl import load_workbook

    workbook_source = io.BytesIO(bytes(source)) if isinstance(source, (bytes, bytearray)) else source
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        actual_sheets = list(workbook.sheetnames)
    finally:
        workbook.close()
    parsed: list[tuple[str, int, int]] = []
    for name in actual_sheets:
        match = re.fullmatch(r"(\d{4})-(0[1-9]|1[0-2])", str(name).strip())
        if not match:
            raise ValueError("往月销量原始表的sheet必须使用YYYY-MM格式")
        parsed.append((str(name).strip(), int(match.group(1)), int(match.group(2))))
    parsed.sort(key=lambda item: (item[1], item[2]))
    if len(parsed) != 12:
        raise ValueError("往月销量原始表必须且只能包含连续12个月的sheet")
    labels = [_month_label(year, month) for _, year, month in parsed]
    expected = [_month_label(*_shift_month(parsed[0][1], parsed[0][2], offset)) for offset in range(12)]
    if labels != expected:
        raise ValueError("往月销量原始表的sheet必须是连续的12个月")
    if require_latest and labels != _rolling_window_labels(today):
        window = _rolling_window_labels(today)
        raise ValueError(f"往月销量原始表必须覆盖最近12个完整月：{window[0]}至{window[-1]}")
    return _build_sales_history_workbook_summary(source, parsed, title="往月销量原始表")


def build_sales_history_monthly_summary(
    sources: Iterable[tuple[str, Path | bytes | bytearray | io.BytesIO]],
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Build the generic history summary from the rolling monthly CSV files."""

    ordered_sources = sorted(
        [(str(month).strip(), source) for month, source in sources],
        key=lambda item: item[0],
    )
    if len(ordered_sources) != 12:
        raise ValueError("往月销量原始表必须包含连续12个月")
    labels = [month for month, _ in ordered_sources]
    parsed = [
        (int(month[:4]), int(month[5:7]))
        for month in labels
        if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month)
    ]
    if len(parsed) != 12 or any(
        (parsed[index][0] * 12 + parsed[index][1]) != (parsed[index - 1][0] * 12 + parsed[index - 1][1] + 1)
        for index in range(1, len(parsed))
    ):
        raise ValueError("往月销量原始表必须是连续的12个月")

    all_asins: set[str] = set()
    site_totals: dict[str, dict[str, float]] = {}
    daily_by_asin: dict[str, dict[date, float]] = {}
    total_rows = 0
    maximum_columns = 0
    month_dates: list[list[date]] = []

    for month, source in ordered_sources:
        year, month_number = int(month[:4]), int(month[5:7])
        frame = read_local_table(source) if isinstance(source, Path) else read_csv_bytes(bytes(source))
        frame = normalize_sales_history_month_source(frame, month)
        day_columns = [
            f"{month_number:02d}-{day:02d}销量"
            for day in range(1, calendar.monthrange(year, month_number)[1] + 1)
        ]
        maximum_columns = max(maximum_columns, len(frame.columns))
        dates = [date(year, month_number, day) for day in range(1, calendar.monthrange(year, month_number)[1] + 1)]
        month_dates.append(dates)
        total_rows += len(frame)
        for _, source_row in frame.iterrows():
            asin = str(source_row["asin"]).strip().upper()
            all_asins.add(asin)
            daily_values = [float(source_row[column]) for column in day_columns]
            country_code = SALES_HISTORY_2025_COUNTRIES.get(str(source_row["国家"] or "").strip())
            if country_code is None:
                continue
            asin_sites = site_totals.setdefault(
                asin,
                {code: 0.0 for code in SALES_HISTORY_2025_COUNTRIES.values()},
            )
            daily_sum = sum(daily_values)
            asin_sites[country_code] += daily_sum
            asin_daily = daily_by_asin.setdefault(asin, {})
            for day, value in zip(dates, daily_values):
                asin_daily[day] = asin_daily.get(day, 0.0) + value

    all_dates = [day for dates in month_dates for day in dates]

    def excluded_stockout_dates(values: list[float]) -> set[date]:
        excluded: set[date] = set()
        run_start: int | None = None
        for index, value in enumerate(values + [1.0]):
            if value == 0 and run_start is None:
                run_start = index
            elif value != 0 and run_start is not None:
                if index - run_start >= 10:
                    excluded.update(all_dates[run_start:index])
                run_start = None
        return excluded

    rows: list[dict[str, object]] = []
    for asin in sorted(all_asins):
        row: dict[str, object] = {"ASIN": asin}
        asin_sites = site_totals.get(asin, {code: 0.0 for code in SALES_HISTORY_2025_COUNTRIES.values()})
        for code in SALES_HISTORY_2025_COUNTRIES.values():
            row[f"{code}总销量"] = _history_output_number(asin_sites[code])
        daily_values = [daily_by_asin.get(asin, {}).get(day, 0.0) for day in all_dates]
        excluded = excluded_stockout_dates(daily_values)
        for index, (label, dates) in enumerate(zip(labels, month_dates), start=1):
            total_sales = sum(daily_by_asin.get(asin, {}).get(day, 0.0) for day in dates)
            included_days = sum(day not in excluded for day in dates)
            adjusted_average = round(total_sales / included_days, 2) if total_sales and included_days else 0.0
            row[f"历史月份{index}"] = label
            row[f"历史{index}月总销量"] = _history_output_number(total_sales)
            row[f"历史{index}月计入天数"] = included_days
            row[f"历史{index}月日均销量"] = adjusted_average
        rows.append(row)
    summary = normalize_sales_history_summary(pd.DataFrame(rows, columns=SALES_HISTORY_GENERIC_COLUMNS))
    return summary, {"rows": total_rows, "effective_rows": len(summary), "columns": maximum_columns}


def normalize_replenishment_gross_profit_source(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in REPLENISHMENT_GROSS_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"毛利原始表缺少补货管理列：{', '.join(missing)}")
    sales_columns = columns_between(df, "销售额--FBA销售额", "COD")

    base = df[["ASIN", "MSKU", "国家", "毛利润"] + GROSS_PROFIT_VOLUME_COLUMNS + sales_columns + REPLENISHMENT_GROSS_RATIO_COLUMNS].copy()
    for col in ["ASIN", "MSKU", "国家"]:
        base[col] = base[col].fillna("").astype(str).str.strip()
    base["ASIN"] = base["ASIN"].str.upper()
    for col in set(["毛利润"] + GROSS_PROFIT_VOLUME_COLUMNS + sales_columns):
        base[col] = normalize_config_number(base[col]).fillna(0)
    for col in REPLENISHMENT_GROSS_RATIO_COLUMNS:
        base[col] = normalize_rate(base[col]).fillna(0)
    base["销量"] = base[GROSS_PROFIT_VOLUME_COLUMNS].sum(axis=1)
    base["销售额"] = base[sales_columns].sum(axis=1)
    return base[
        ["ASIN", "MSKU", "国家", "销量", "销售额", "毛利润"] + REPLENISHMENT_GROSS_RATIO_COLUMNS
    ]


def build_replenishment_gross_summary(df: pd.DataFrame) -> pd.DataFrame:
    gross_profit = normalize_replenishment_gross_profit_source(df)
    if gross_profit.empty:
        return pd.DataFrame(columns=["ASIN"] + replenishment_gross_columns())

    country_data = gross_profit[gross_profit["国家"].isin(PRODUCT_COUNTRIES)].copy()
    if country_data.empty:
        return pd.DataFrame(columns=["ASIN"] + replenishment_gross_columns())

    grouped = (
        country_data.groupby(["ASIN", "国家"], dropna=False, sort=False)
        .agg(单量=("销量", "sum"), 销售额=("销售额", "sum"), 毛利润=("毛利润", "sum"))
        .reset_index()
    )
    grouped["毛利率"] = grouped.apply(lambda row: safe_blank_ratio(row["毛利润"], row["销售额"]), axis=1)
    result = grouped[["ASIN"]].drop_duplicates().copy()

    for country in PRODUCT_COUNTRIES:
        subset = grouped[grouped["国家"].eq(country)][["ASIN", "单量", "毛利率"]].copy()
        subset = subset.rename(columns={"单量": f"{country}单量", "毛利率": f"{country}毛利率"})
        result = result.merge(subset, on="ASIN", how="left")

    reason_wide = build_replenishment_reason_columns(country_data)
    result = result.merge(reason_wide, on="ASIN", how="left")
    for col in replenishment_gross_columns():
        if col not in result.columns:
            result[col] = pd.NA if not col.endswith("原因") else ""
    return result[["ASIN"] + replenishment_gross_columns()]


def build_replenishment_reason_columns(gross_profit: pd.DataFrame) -> pd.DataFrame:
    if gross_profit.empty:
        return pd.DataFrame(columns=["ASIN"] + replenishment_reason_columns())

    rows = []
    grouped = gross_profit.groupby(["ASIN", "国家", "MSKU"], dropna=False, sort=False)
    for (asin, country, msku), group in grouped:
        if country not in PRODUCT_COUNTRIES or not str(msku).strip():
            continue
        reasons = []
        if group["广告费占比"].max() > 0.15:
            reasons.append("广告炸")
        if group["退款占比"].max() > 0.08:
            reasons.append("退货多")
        if group["FBA发货费占比"].max() > 0.60:
            reasons.append("FBA配送费炸")
        if reasons:
            rows.append({"ASIN": asin, "国家": country, "MSKU原因": f"{msku}: {'、'.join(reasons)}"})

    if not rows:
        return pd.DataFrame(columns=["ASIN"] + replenishment_reason_columns())

    reason_data = pd.DataFrame(rows)
    grouped_reasons = (
        reason_data.groupby(["ASIN", "国家"], dropna=False, sort=False)
        .agg(原因=("MSKU原因", lambda values: "；".join(values)))
        .reset_index()
    )
    result = grouped_reasons[["ASIN"]].drop_duplicates().copy()
    for country in PRODUCT_COUNTRIES:
        subset = grouped_reasons[grouped_reasons["国家"].eq(country)][["ASIN", "原因"]].copy()
        subset = subset.rename(columns={"原因": f"{country}原因"})
        result = result.merge(subset, on="ASIN", how="left")
    for col in replenishment_reason_columns():
        if col not in result.columns:
            result[col] = ""
    return result[["ASIN"] + replenishment_reason_columns()]


def replenishment_gross_columns() -> list[str]:
    columns = []
    for country in PRODUCT_COUNTRIES:
        columns.extend([f"{country}单量", f"{country}毛利率", f"{country}原因"])
    return columns


def replenishment_reason_columns() -> list[str]:
    return [f"{country}原因" for country in PRODUCT_COUNTRIES]


def build_replenishment_rating_summary(df: pd.DataFrame) -> pd.DataFrame:
    rating = normalize_rating_source(df)
    if rating.empty:
        return pd.DataFrame(columns=["ASIN", "产品评价数", "产品评分值"])

    rating["ASIN"] = rating["ASIN"].fillna("").astype(str).str.strip().str.upper()
    data = rating[rating["国家"].isin(PRODUCT_COUNTRIES)].copy()
    if data.empty:
        return pd.DataFrame(columns=["ASIN", "产品评价数", "产品评分值"])
    country_order = {country: index for index, country in enumerate(PRODUCT_COUNTRIES)}
    data["_国家排序"] = data["国家"].map(country_order).fillna(len(country_order))
    data = data.sort_values(["ASIN", "Rating总数", "_国家排序"], ascending=[True, False, True], kind="stable")
    best = data.drop_duplicates(subset=["ASIN"], keep="first").copy()
    best["产品评价数"] = pd.to_numeric(best["Rating总数"], errors="coerce")
    best["产品评分值"] = pd.to_numeric(best["评分"], errors="coerce")
    return best[["ASIN", "产品评价数", "产品评分值"]].reset_index(drop=True)


def build_replenishment_management_tables(
    operational_df: pd.DataFrame,
    gross_profit_df: pd.DataFrame | None = None,
    rating_df: pd.DataFrame | None = None,
    target_config: pd.DataFrame | None = None,
    *,
    coverage_rules: pd.DataFrame | None = None,
    replenishment_switches: pd.DataFrame | None = None,
    product_tags: pd.DataFrame | None = None,
    store_config: pd.DataFrame | None = None,
    sales_history_2025: pd.DataFrame | None = None,
    sales_history_rolling: pd.DataFrame | None = None,
    promotions: pd.DataFrame | None = None,
    only_needed: bool = True,
    today: date | datetime | pd.Timestamp | None = None,
) -> dict[str, pd.DataFrame]:
    """Build Excel-compatible, ASIN-group replenishment recommendations.

    ``target_config`` remains accepted so older integrations keep calling the
    function successfully, but target days and case packs intentionally no
    longer participate in the calculation.
    """
    del target_config
    operational = normalize_replenishment_operational(operational_df)
    rules = normalize_replenishment_coverage_rules(coverage_rules)
    if rules.empty:
        rules = default_replenishment_coverage_rules()
    switches = normalize_replenishment_switches(replenishment_switches)
    tags = normalize_replenishment_product_tags(product_tags)
    stores = normalize_store_config(store_config if store_config is not None else pd.DataFrame())
    if sales_history_rolling is not None and not sales_history_rolling.empty:
        history = normalize_sales_history_summary(sales_history_rolling)
    elif sales_history_2025 is not None and not sales_history_2025.empty:
        history = _legacy_to_generic_history(sales_history_2025)
    else:
        history = pd.DataFrame(columns=SALES_HISTORY_GENERIC_COLUMNS)

    if operational.empty:
        return {
            "detail": pd.DataFrame(columns=replenishment_management_columns()),
            "sku_detail": pd.DataFrame(columns=replenishment_sku_detail_columns()),
            "history": history,
        }

    data = operational.copy()
    data["补货组ID"] = data.apply(
        lambda row: str(row["ASIN"])
        if str(row["ASIN"]).strip()
        else f"异常-{str(row['MSKU']).strip() or row.name + 1}",
        axis=1,
    )
    calculation_day = pd.Timestamp(today).normalize() if today is not None else pd.Timestamp.today().normalize()
    data["上架天数"] = (calculation_day - data["上架时间"].dt.normalize()).dt.days
    numeric_required = REPLENISHMENT_SALES_COLUMNS + REPLENISHMENT_FORMULA_STOCK_COLUMNS + ["单品重量(g)"]
    data["数据异常"] = ""
    for index, row in data.iterrows():
        issues: list[str] = []
        if not str(row["ASIN"]).strip():
            issues.append("缺少ASIN")
        if not str(row["MSKU"]).strip():
            issues.append("缺少MSKU")
        if pd.isna(row["上架时间"]):
            issues.append("缺少上架时间")
        elif pd.isna(row["上架天数"]) or row["上架天数"] < 0:
            issues.append("上架时间无效")
        for column in numeric_required:
            if pd.isna(row[column]):
                issues.append(f"缺少{column}")
        data.at[index, "数据异常"] = "；".join(issues)

    valid = data["数据异常"].eq("")
    data["校准日销量"] = pd.NA
    new_sku = valid & data["上架天数"].lt(90)
    old_sku = valid & ~new_sku
    data.loc[new_sku, "校准日销量"] = (
        data.loc[new_sku, "7天销量"] / 7 * 0.90
        + data.loc[new_sku, "14天销量"] / 14 * 0.09
        + data.loc[new_sku, "30天销量"] / 30 * 0.01
    )
    data.loc[old_sku, "校准日销量"] = (
        data.loc[old_sku, "7天销量"] / 7 * 0.60
        + data.loc[old_sku, "14天销量"] / 14 * 0.30
        + data.loc[old_sku, "30天销量"] / 30 * 0.10
    )
    data["SKU总库存"] = (
        data["可售"] + data["待调仓"] + data["调仓中"] + data["待入库"] * 2
        + data["采购在途"] + data["本地库存"] + data["在途"] + data["计划入库"]
    )
    data["SKU亚马逊可售"] = data[["可售", "待调仓", "调仓中", "待入库"]].sum(axis=1, min_count=4)
    data["T值"] = data["7天销量"] / 7 - data["30天销量"] / 30
    data["库龄90天以上"] = data[AGING_STOCK_COLUMNS].sum(axis=1, min_count=1)
    data["库龄180-365天"] = data[["181-330天库存数", "331-365天库存数"]].sum(axis=1, min_count=1)
    data["库龄365天以上"] = data[["366-455天库存数", "456天以上库存数"]].sum(axis=1, min_count=1)
    stopped_codes = {
        extract_store_code(value)
        for value in stores.loc[stores["停提款时间"].fillna("").astype(str).str.strip().ne(""), "店铺名"]
        if extract_store_code(value)
    }

    def store_status_text(value: object) -> str:
        codes = [code for code, _ in extract_operational_store_codes(value) if code]
        if not codes:
            return str(value or "").strip()
        return "；".join(f"{code}·{'停运' if code in stopped_codes else '正常'}" for code in dict.fromkeys(codes))

    def join_store_statuses(values: pd.Series) -> str:
        parts: list[str] = []
        seen: set[str] = set()
        for value in values:
            for part in str(value or "").split("；"):
                part = part.strip()
                if part and part not in seen:
                    parts.append(part)
                    seen.add(part)
        return "；".join(parts)

    data["店铺状态"] = data["店铺名称"].map(store_status_text)
    data["SKU角色"] = "跟卖SKU"
    asin_inventory = data.groupby("ASIN", dropna=False, sort=False)["SKU总库存"].agg(
        lambda values: values.sum(min_count=len(values))
    )

    switch_map = {
        str(row["ASIN"]): (bool(row["是否补货"]), str(row["关闭原因"]))
        for _, row in switches.iterrows()
    }
    parent_rows: list[dict] = []
    sku_rows: list[dict] = []
    enabled_rules = rules[rules["是否启用"]].copy()
    enabled_rules["重量上限排序"] = enabled_rules["重量上限"].fillna(float("inf"))
    enabled_rules = enabled_rules.sort_values(["重量下限", "重量上限排序"], kind="stable")

    for group_id, group in data.groupby("补货组ID", sort=False, dropna=False):
        group = group.copy()
        group_issues = [item for item in group["数据异常"].tolist() if item]
        original_index = group.sort_values(["上架时间", "MSKU"], kind="stable", na_position="last").index[0]
        group.loc[original_index, "SKU角色"] = "原SKU"
        max_weight = group["单品重量(g)"].max(skipna=True)
        rule = pd.DataFrame()
        if not group_issues and pd.notna(max_weight):
            rule = enabled_rules[(enabled_rules["重量下限"] <= max_weight) & (enabled_rules["重量上限"].isna() | (enabled_rules["重量上限"] > max_weight))]
            if len(rule) != 1:
                group_issues.append("未匹配唯一库存覆盖规则")
        coverage_days = pd.NA
        transportation = ""
        if len(rule) == 1:
            matched = rule.iloc[0]
            transportation = str(matched["运输方式"])
            coverage_days = int(matched["头程时效"] + matched["预警天数"] + matched["补货频次"])
        calibrated = group["校准日销量"].sum(min_count=len(group)) if not group_issues else pd.NA
        inventory = group["SKU总库存"].sum(min_count=len(group)) if not group_issues else pd.NA
        amazon_available = group["SKU亚马逊可售"].sum(min_count=len(group))
        trend = group["T值"].sum(min_count=len(group))
        asin_reference_inventory = sum(
            (asin_inventory.get(asin, pd.NA) for asin in group["ASIN"].dropna().astype(str).unique()),
            start=0,
        )
        if any(pd.isna(asin_inventory.get(asin, pd.NA)) for asin in group["ASIN"].dropna().astype(str).unique()):
            asin_reference_inventory = pd.NA
        target_stock = calibrated * coverage_days if not group_issues else pd.NA
        measured = excel_round_to_ten(target_stock - inventory) if not group_issues else pd.NA
        replenishment_enabled, close_reason = switch_map.get(str(group_id), (True, ""))
        official = measured if replenishment_enabled and not group_issues else (0 if not group_issues else pd.NA)
        status = "数据异常" if group_issues else ("已关闭补货" if not replenishment_enabled else "正常")
        asin_values = join_non_empty_values(group["ASIN"])
        follower_skus = join_non_empty_values(group.loc[group.index != original_index, "MSKU"])
        parent_rows.append(
            {
                "补货组ID": str(group_id), "ASIN": asin_values, "原SKU": str(group.loc[original_index, "MSKU"]),
                "跟卖SKU": follower_skus, "SKU数量": int(group["MSKU"].nunique()), "店铺编码": join_operational_store_codes(group["店铺名称"]),
                "店铺状态": join_store_statuses(group["店铺状态"]),
                "开发员": join_non_empty_values(group["开发员"]), "最大重量(g)": max_weight,
                "库存覆盖天数": coverage_days, "T值": trend, "校准日销量": calibrated,
                "亚马逊可售": amazon_available, "总可售": inventory, "跟卖总可售": asin_reference_inventory,
                "ASIN总库存": inventory,
                "库龄90天以上": group["库龄90天以上"].sum(min_count=len(group)),
                "库龄180-365天": group["库龄180-365天"].sum(min_count=len(group)),
                "库龄365天以上": group["库龄365天以上"].sum(min_count=len(group)),
                "目标库存": target_stock, "测算建议补货数量": measured, "建议补货数量": official,
                "是否补货": replenishment_enabled, "关闭原因": close_reason,
                "数据状态": status, "数据异常": "；".join(dict.fromkeys(group_issues)),
                "运输方式": transportation,
            }
        )
        for _, row in group.sort_values(["上架时间", "MSKU"], kind="stable", na_position="last").iterrows():
            sku_rows.append(
                {
                    "补货组ID": str(group_id), "ASIN": row["ASIN"], "SKU角色": row["SKU角色"], "MSKU": row["MSKU"],
                    "店铺名称": row["店铺名称"], "店铺状态": row["店铺状态"], "开发员": row["开发员"],
                    "上架时间": row["上架时间"], "上架天数": row["上架天数"],
                    "单品重量(g)": row["单品重量(g)"], "7天销量": row["7天销量"], "14天销量": row["14天销量"],
                    "30天销量": row["30天销量"], "T值": row["T值"], "校准日销量": row["校准日销量"],
                    "SKU亚马逊可售": row["SKU亚马逊可售"], "SKU总库存": row["SKU总库存"],
                    "库龄90天以上": row["库龄90天以上"], "库龄180-365天": row["库龄180-365天"],
                    "库龄365天以上": row["库龄365天以上"],
                    **{column: row[column] for column in REPLENISHMENT_FORMULA_STOCK_COLUMNS},
                    "数据异常": row["数据异常"],
                }
            )

    detail = pd.DataFrame(parent_rows)
    sku_detail = pd.DataFrame(sku_rows)
    if promotions is not None and not promotions.empty and not sku_detail.empty:
        # These optional output columns are part of the canonical schema and
        # therefore start as NA.  Remove the placeholders before merging so a
        # matching promotion does not become pandas' _x/_y duplicate columns.
        sku_detail = sku_detail.drop(
            columns=[
                "最近促销开始日期", "最近促销截止日期", "最近促销折扣",
                "promotion_name", "promotion_updated_at",
            ],
            errors="ignore",
        )
        promo = promotions.copy()
        promo = promo.rename(
            columns={
                "sku": "MSKU",
                "start_date": "最近促销开始日期",
                "end_date": "最近促销截止日期",
                "discount_percent": "最近促销折扣",
                "updated_at": "promotion_updated_at",
            }
        )
        keep = [
            column
            for column in [
                "MSKU", "最近促销开始日期", "最近促销截止日期", "最近促销折扣",
                "promotion_name", "promotion_updated_at",
            ]
            if column in promo.columns
        ]
        sku_detail = sku_detail.merge(promo[keep].drop_duplicates("MSKU", keep="last"), on="MSKU", how="left")
    for column in [
        "最近促销开始日期", "最近促销截止日期", "最近促销折扣",
        "promotion_name", "promotion_updated_at",
    ]:
        if column not in sku_detail.columns:
            sku_detail[column] = pd.NA
    if not sku_detail.empty:
        promotion_rows: list[dict[str, object]] = []
        for group_id, group in sku_detail.groupby("补货组ID", sort=False, dropna=False):
            candidates = group[
                group[
                    [
                        "最近促销开始日期", "最近促销截止日期", "最近促销折扣",
                        "promotion_name", "promotion_updated_at",
                    ]
                ]
                .notna()
                .any(axis=1)
            ].copy()
            if candidates.empty:
                continue
            candidates["_开始日期"] = pd.to_datetime(candidates["最近促销开始日期"], errors="coerce")
            candidates["_截止日期"] = pd.to_datetime(candidates["最近促销截止日期"], errors="coerce")
            candidates["_更新时间"] = pd.to_datetime(candidates["promotion_updated_at"], errors="coerce")
            active = candidates[
                candidates["_开始日期"].le(calculation_day)
                & (candidates["_截止日期"].isna() | candidates["_截止日期"].ge(calculation_day))
            ]
            if not active.empty:
                selected = active.sort_values(
                    ["_截止日期", "_开始日期", "_更新时间"],
                    ascending=[True, False, False],
                    na_position="last",
                    kind="stable",
                ).iloc[0]
            else:
                ended = candidates[candidates["_截止日期"].lt(calculation_day)]
                if ended.empty:
                    continue
                selected = ended.sort_values(
                    ["_截止日期", "_开始日期", "_更新时间"],
                    ascending=[False, False, False],
                    na_position="last",
                    kind="stable",
                ).iloc[0]
            promotion_rows.append(
                {
                    "补货组ID": str(group_id),
                    "最近促销开始日期": selected.get("最近促销开始日期"),
                    "最近促销截止日期": selected.get("最近促销截止日期"),
                    "最近促销折扣": selected.get("最近促销折扣"),
                }
            )
        if promotion_rows:
            detail = detail.merge(pd.DataFrame(promotion_rows), on="补货组ID", how="left")

    enabled_tags = tags[tags["是否启用"]].copy()
    if not enabled_tags.empty:
        tag_summary = (
            enabled_tags.groupby("ASIN", sort=False, dropna=False)
            .agg(
                产品标签=("产品标签", join_non_empty_values),
                产品标签颜色=("标签颜色", lambda values: "；".join(str(value).strip() for value in values)),
            )
            .reset_index()
        )
        detail = detail.merge(tag_summary, on="ASIN", how="left")
        sku_detail = sku_detail.merge(tag_summary, on="ASIN", how="left")

    if not history.empty:
        detail = detail.merge(history[SALES_HISTORY_GENERIC_COLUMNS], on="ASIN", how="left")
    for column in replenishment_sku_detail_columns():
        if column not in sku_detail.columns:
            sku_detail[column] = pd.NA
    sku_detail = sku_detail[replenishment_sku_detail_columns()]
    if gross_profit_df is not None and not gross_profit_df.empty and not detail.empty:
        detail = detail.merge(build_replenishment_gross_summary(gross_profit_df), on="ASIN", how="left")
    if rating_df is not None and not rating_df.empty and not detail.empty:
        detail = detail.merge(build_replenishment_rating_summary(rating_df), on="ASIN", how="left")
    for column in replenishment_management_columns():
        if column not in detail.columns:
            detail[column] = pd.NA
    detail = detail[replenishment_management_columns()]
    if only_needed:
        detail = detail[detail["建议补货数量"].fillna(0).gt(0) | detail["数据状态"].eq("数据异常")].copy()
        sku_detail = sku_detail[sku_detail["补货组ID"].isin(set(detail["补货组ID"]))].copy()
    detail = detail.sort_values(["建议补货数量", "ASIN", "补货组ID"], ascending=[False, True, True], kind="stable", na_position="last").reset_index(drop=True)
    return {
        "detail": detail,
        "sku_detail": sku_detail.reset_index(drop=True),
        "history": _history_with_legacy_aliases(history),
    }


def build_replenishment_store_distribution(detail: pd.DataFrame) -> pd.DataFrame:
    # Recommendations are product-level purchase quantities and are not
    # allocated to stores in this version.
    return pd.DataFrame(columns=["店铺编码", "需补货ASIN数"])


def replenishment_management_columns() -> list[str]:
    return [
        "补货组ID", "ASIN", "原SKU", "跟卖SKU", "SKU数量", "店铺编码", "店铺状态", "开发员",
        "最大重量(g)", "库存覆盖天数", "T值", "校准日销量", "亚马逊可售", "总可售",
        "跟卖总可售", "ASIN总库存", "库龄90天以上", "库龄180-365天", "库龄365天以上",
        "目标库存", "测算建议补货数量", "建议补货数量", "是否补货", "关闭原因",
        "数据状态", "数据异常", "最近促销开始日期", "最近促销截止日期", "最近促销折扣",
        "产品标签", "产品标签颜色", *SALES_HISTORY_GENERIC_COLUMNS[1:],
    ] + replenishment_gross_columns() + ["产品评价数", "产品评分值"]


def replenishment_sku_detail_columns() -> list[str]:
    return [
        "补货组ID", "ASIN", "SKU角色", "MSKU", "店铺名称", "店铺状态", "开发员",
        "上架时间", "上架天数", "单品重量(g)", "7天销量", "14天销量", "30天销量",
        "T值", "校准日销量", "SKU亚马逊可售", "SKU总库存", "库龄90天以上",
        "库龄180-365天", "库龄365天以上", *REPLENISHMENT_FORMULA_STOCK_COLUMNS,
        "最近促销开始日期", "最近促销截止日期", "最近促销折扣", "promotion_name",
        "产品标签", "产品标签颜色", "数据异常",
    ]


def excel_round_to_ten(gap: float | int | pd.NA) -> int:
    """Excel ROUND(gap/10, 0) * 10, then MAX(..., 0)."""
    if pd.isna(gap) or float(gap) <= 0:
        return 0
    # Weighted daily sales can produce values such as 34.99999999999999 even
    # when the business value is exactly 35. Excel rounds that up, so retain a
    # tiny numerical tolerance before applying the half-up rule.
    return int(math.floor(float(gap) / 10 + 0.5 + 1e-9) * 10)


def normalize_product_operational(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in PRODUCT_OPERATIONAL_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"运营原始表缺少产品管理列：{', '.join(missing)}")

    base = df[PRODUCT_OPERATIONAL_REQUIRED_COLUMNS].copy()
    for col in ["ASIN", "MSKU"]:
        base[col] = base[col].fillna("").astype(str).str.strip()
    base = base[base["MSKU"].ne("")].copy()
    for col in PRODUCT_OPERATIONAL_REQUIRED_COLUMNS:
        if col in {"ASIN", "MSKU"}:
            continue
        base[col] = normalize_config_number(base[col]).fillna(0)
    base = base.rename(columns={"MSKU": "SKU", "可售": "可售数量"})
    if base.empty:
        return pd.DataFrame(columns=["ASIN", "SKU", "可售天数"] + PRODUCT_OPERATIONAL_SUM_COLUMNS)

    grouped = (
        base.groupby(["ASIN", "SKU"], dropna=False, sort=False)
        .agg(
            可售数量=("可售数量", "sum"),
            可售天数=("可售天数", "first"),
            日均销量=("日均销量", "sum"),
            昨天销量=("昨天销量", "sum"),
            前天销量=("前天销量", "sum"),
            上前销量=("上前销量", "sum"),
            **{
                "7天销量": ("7天销量", "sum"),
                "14天销量": ("14天销量", "sum"),
                "30天销量": ("30天销量", "sum"),
                "90天销量": ("90天销量", "sum"),
            },
        )
        .reset_index()
    )
    return grouped[["ASIN", "SKU", "可售数量", "可售天数"] + [col for col in PRODUCT_OPERATIONAL_SUM_COLUMNS if col != "可售数量"]]


def normalize_gross_profit_source(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in GROSS_PROFIT_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"毛利原始表缺少列：{', '.join(missing)}")
    sales_columns = columns_between(df, "销售额--FBA销售额", "COD")

    base = df[["ASIN", "MSKU", "国家", "毛利润"] + GROSS_PROFIT_VOLUME_COLUMNS + GROSS_PROFIT_AD_COLUMNS + sales_columns].copy()
    for col in ["ASIN", "MSKU", "国家"]:
        base[col] = base[col].fillna("").astype(str).str.strip()
    for col in set(["毛利润"] + GROSS_PROFIT_VOLUME_COLUMNS + GROSS_PROFIT_AD_COLUMNS + sales_columns):
        base[col] = normalize_config_number(base[col]).fillna(0)
    base["销量"] = base[GROSS_PROFIT_VOLUME_COLUMNS].sum(axis=1)
    base["销售额"] = base[sales_columns].sum(axis=1)
    base["广告支出净额"] = base[GROSS_PROFIT_AD_COLUMNS].sum(axis=1)
    return base[["ASIN", "MSKU", "国家", "销量", "销售额", "毛利润", "广告支出净额"]]


def normalize_rating_source(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in RATING_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Rating缺少列：{', '.join(missing)}")
    base = df[RATING_REQUIRED_COLUMNS].copy()
    for col in ["ASIN", "国家"]:
        base[col] = base[col].fillna("").astype(str).str.strip()
    base["Rating总数"] = normalize_config_number(base["Rating总数"]).fillna(0)
    base["评分"] = normalize_config_number(base["评分"])
    return base


def build_product_management_table(
    operational_df: pd.DataFrame, gross_profit_df: pd.DataFrame, rating_df: pd.DataFrame
) -> pd.DataFrame:
    operational = normalize_product_operational(operational_df)
    gross_profit = normalize_gross_profit_source(gross_profit_df)
    rating = normalize_rating_source(rating_df)
    if operational.empty:
        return pd.DataFrame(columns=product_management_all_columns())

    result = operational.copy()
    result["_sort_order"] = range(len(result))
    sku_gross = build_product_gross_columns(gross_profit, ["ASIN", "MSKU"]).rename(columns={"MSKU": "SKU"})
    rating_wide = build_product_rating_columns(rating)

    result = result.merge(sku_gross, on=["ASIN", "SKU"], how="left").merge(rating_wide, on="ASIN", how="left")
    for col in product_management_all_columns():
        if col not in result.columns:
            result[col] = pd.NA
    return result[product_management_all_columns()].sort_values("_sort_order", kind="stable").reset_index(drop=True)


def build_low_margin_product_table(
    gross_profit_df: pd.DataFrame,
    threshold: float = LOW_MARGIN_PRODUCT_THRESHOLD,
    min_sales: float = LOW_MARGIN_PRODUCT_MIN_SALES,
    developers: Iterable[str] | None = None,
    allowed_skus: Iterable[str] | None = None,
) -> pd.DataFrame:
    gross_profit = normalize_low_margin_gross_profit_source(gross_profit_df)
    if allowed_skus is not None:
        normalized_allowed_skus = {
            str(sku).strip().upper()
            for sku in allowed_skus
            if str(sku).strip()
        }
        normalized_skus = gross_profit["MSKU"].fillna("").astype(str).str.strip().str.upper()
        gross_profit = gross_profit[normalized_skus.isin(normalized_allowed_skus)].copy()
    if developers:
        selected_developers = {str(developer).strip() for developer in developers if str(developer).strip()}
        gross_profit = gross_profit[gross_profit["开发员"].isin(selected_developers)].copy()
    if gross_profit.empty:
        return pd.DataFrame(columns=LOW_MARGIN_PRODUCT_COLUMNS)

    grouped = (
        gross_profit.rename(columns={"MSKU": "SKU"})
        .groupby(["ASIN", "SKU", "国家"], dropna=False, sort=False)
        .agg(
            开发员=("开发员", join_non_empty_values),
            销量=("销量", "sum"),
            销售额=("销售额", "sum"),
            毛利润=("毛利润", "sum"),
        )
        .reset_index()
    )
    grouped["毛利率"] = grouped.apply(lambda row: safe_blank_ratio(row["毛利润"], row["销售额"]), axis=1)
    grouped = grouped[
        (pd.to_numeric(grouped["销量"], errors="coerce") >= min_sales)
        & (pd.to_numeric(grouped["毛利率"], errors="coerce") < threshold)
    ].copy()

    if grouped.empty:
        return pd.DataFrame(columns=LOW_MARGIN_PRODUCT_COLUMNS)

    grouped = grouped.sort_values(["毛利率", "SKU", "国家"], ascending=[True, True, True], kind="stable")
    return grouped[LOW_MARGIN_PRODUCT_COLUMNS].reset_index(drop=True)


def normalize_low_margin_gross_profit_source(df: pd.DataFrame) -> pd.DataFrame:
    gross_profit = normalize_gross_profit_source(df)
    developer_col = first_existing_column(df, GROSS_PROFIT_DEVELOPER_COLUMNS)
    if developer_col is None:
        gross_profit["开发员"] = ""
        return gross_profit

    gross_profit["开发员"] = df[developer_col].fillna("").astype(str).str.strip().reset_index(drop=True)
    return gross_profit


def first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for col in candidates:
        if col in df.columns:
            return col
    return None


def join_non_empty_values(values: pd.Series) -> str:
    return "；".join(sorted({str(value).strip() for value in values if str(value).strip()}))


def sort_key_series(series: pd.Series) -> pd.Series:
    non_empty = series.notna() & series.astype(str).str.strip().ne("")
    numeric = normalize_config_number(series)
    if non_empty.any() and numeric.notna().sum() / non_empty.sum() >= 0.8:
        return numeric
    return series.map(normalize_sort_text)


def normalize_sort_text(value) -> str:
    if pd.isna(value):
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip().casefold()


def sort_product_management_table(df: pd.DataFrame, sort_column: str | None = None, ascending: bool = True) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    result = df.copy()
    if not sort_column or sort_column == "默认排序" or sort_column not in result.columns:
        if "_sort_order" in result.columns:
            return result.sort_values("_sort_order", kind="stable").reset_index(drop=True)
        return result.reset_index(drop=True)
    order_col = "_sort_order" if "_sort_order" in result.columns else sort_column
    return result.sort_values(
        [sort_column, order_col],
        ascending=[ascending, True],
        na_position="last",
        kind="stable",
        key=sort_key_series,
    ).reset_index(drop=True)


def build_product_gross_columns(gross_profit: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if gross_profit.empty:
        return pd.DataFrame(columns=group_cols + product_gross_columns())

    totals = (
        gross_profit.groupby(group_cols, dropna=False, sort=False)
        .agg(销售额=("销售额", "sum"), 毛利润=("毛利润", "sum"))
        .reset_index()
    )
    totals["毛利率"] = totals.apply(lambda row: safe_blank_ratio(row["毛利润"], row["销售额"]), axis=1)

    country_data = gross_profit[gross_profit["国家"].isin(PRODUCT_COUNTRIES)].copy()
    if country_data.empty:
        for col in product_country_gross_columns():
            totals[col] = pd.NA
        return totals[group_cols + product_gross_columns()]

    country_grouped = (
        country_data.groupby(group_cols + ["国家"], dropna=False, sort=False)
        .agg(销量=("销量", "sum"), 销售额=("销售额", "sum"), 毛利润=("毛利润", "sum"), 广告支出净额=("广告支出净额", "sum"))
        .reset_index()
    )
    country_grouped["毛利率"] = country_grouped.apply(lambda row: safe_blank_ratio(row["毛利润"], row["销售额"]), axis=1)
    country_grouped["广告费占比"] = country_grouped.apply(
        lambda row: safe_blank_ratio(abs(row["广告支出净额"]), row["销售额"]),
        axis=1,
    )

    wide = totals
    for country in PRODUCT_COUNTRIES:
        subset = country_grouped[country_grouped["国家"].eq(country)][group_cols + ["销量", "毛利率", "广告费占比"]].copy()
        subset = subset.rename(
            columns={
                "销量": f"{country}销量",
                "毛利率": f"{country}毛利率",
                "广告费占比": f"{country}广告费占比",
            }
        )
        wide = wide.merge(subset, on=group_cols, how="left")
    for col in product_gross_columns():
        if col not in wide.columns:
            wide[col] = pd.NA
    return wide[group_cols + product_gross_columns()]


def build_product_rating_columns(rating: pd.DataFrame) -> pd.DataFrame:
    if rating.empty:
        return pd.DataFrame(columns=["ASIN"] + product_rating_columns())
    base = rating[rating["国家"].isin(PRODUCT_COUNTRIES)].copy()
    if base.empty:
        return pd.DataFrame(columns=["ASIN"] + product_rating_columns())
    grouped = (
        base.groupby(["ASIN", "国家"], dropna=False, sort=False)
        .agg(Rating总数=("Rating总数", "max"), 评分=("评分", "mean"))
        .reset_index()
    )
    country_order = {country: idx for idx, country in enumerate(PRODUCT_COUNTRIES)}
    grouped["_国家排序"] = grouped["国家"].map(country_order).fillna(len(country_order))
    best = grouped.sort_values(
        ["ASIN", "Rating总数", "_国家排序"],
        ascending=[True, False, True],
        kind="stable",
    ).drop_duplicates("ASIN", keep="first")
    best["Rating"] = best.apply(format_product_rating, axis=1)
    return best[["ASIN"] + product_rating_columns()].reset_index(drop=True)


def columns_between(df: pd.DataFrame, start: str, end: str) -> list[str]:
    columns = list(df.columns)
    if start not in columns or end not in columns:
        missing = [col for col in [start, end] if col not in columns]
        raise ValueError(f"毛利原始表缺少销售额区间列：{', '.join(missing)}")
    start_index = columns.index(start)
    end_index = columns.index(end)
    if start_index > end_index:
        raise ValueError(f"毛利原始表销售额区间列顺序异常：{start} 在 {end} 之后")
    return columns[start_index : end_index + 1]


def safe_blank_ratio(numerator, denominator):
    return numerator / denominator if pd.notna(denominator) and denominator else pd.NA


def format_product_rating(row) -> str:
    count = row["Rating总数"]
    score = row["评分"]
    if pd.isna(count):
        return ""
    count_text = str(int(round(float(count))))
    if pd.isna(score):
        return count_text
    return f"{count_text}({float(score):.1f})"


def product_country_gross_columns() -> list[str]:
    columns = []
    for country in PRODUCT_COUNTRIES:
        columns.extend([f"{country}销量", f"{country}毛利率", f"{country}广告费占比"])
    return columns


def product_gross_columns() -> list[str]:
    return product_country_gross_columns() + ["销售额", "毛利润", "毛利率"]


def product_rating_columns() -> list[str]:
    return ["Rating"]


def product_management_columns() -> list[str]:
    return [
        "SKU",
        "ASIN",
        "可售数量",
        "可售天数",
        "日均销量",
        "昨天销量",
        "前天销量",
        "上前销量",
        "7天销量",
        "14天销量",
        "30天销量",
        "90天销量",
    ] + product_gross_columns() + product_rating_columns()


def product_management_internal_columns() -> list[str]:
    return ["_sort_order"]


def product_management_all_columns() -> list[str]:
    return product_management_columns() + product_management_internal_columns()


def product_management_display_table(df: pd.DataFrame) -> pd.DataFrame:
    display_columns = [col for col in product_management_columns() if col in df.columns]
    return df[display_columns].copy()


def maybe_numeric(series: pd.Series) -> pd.Series:
    if series.dtype.kind in "biufc":
        return series
    text = normalized_numeric_text(series)
    percent_mask = text.str.endswith("%", na=False)
    cleaned = text.str.replace("%", "", regex=False)
    numeric = pd.to_numeric(cleaned, errors="coerce").astype("float64")
    non_empty = (text.notna() & text.ne("")).sum()
    parsed = numeric.notna().sum()
    if non_empty and parsed / non_empty >= 0.8:
        numeric.loc[percent_mask] = numeric.loc[percent_mask] / 100
        return numeric
    return series


def is_enabled(value) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用"}


def merge_business_config(df: pd.DataFrame, store_config: pd.DataFrame, target_config: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()

    store_config = store_config.copy()
    store_config = normalize_store_config(store_config)
    store_config["店铺编码"] = store_config["店铺名"].map(extract_store_code)
    store_config = store_config.rename(columns={"店铺所属部门": "部门"})
    if not store_config.empty:
        result = result.merge(store_config[["店铺编码", "店铺类型", "停提款时间", "部门"]], on="店铺编码", how="left")

    for col in ["店铺类型", "停提款时间", "部门"]:
        if col not in result.columns:
            result[col] = None
    result[["店铺类型", "部门"]] = result[["店铺类型", "部门"]].fillna("未配置")
    result["停提款时间"] = result["停提款时间"].fillna("").map(lambda value: normalize_month(value) or "")
    result["是否停提款数据"] = result.apply(
        lambda row: bool(row["停提款时间"]) and str(row["月份"]) >= str(row["停提款时间"]),
        axis=1,
    )

    target_config = target_config.copy()
    target_config = normalize_target_config(target_config)
    if not target_config.empty:
        target_config = target_config.rename(
            columns={"开发员": "销售专员", "目标业绩": "销售额目标", "目标毛利率": "毛利率目标"}
        )
        target_config["销售额目标"] = maybe_numeric(target_config["销售额目标"])
        target_config["毛利率目标"] = normalize_rate(target_config["毛利率目标"])
        result = result.merge(target_config, on=["销售专员"], how="left")

    for col in ["销售额目标", "毛利率目标"]:
        if col not in result.columns:
            result[col] = pd.NA
    return result


def split_counted_and_stopped_data(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if data.empty:
        return data.copy(), data.copy()
    if "是否停提款数据" not in data.columns:
        return data.copy(), data.iloc[0:0].copy()
    stopped_mask = data["是否停提款数据"].fillna(False).astype(bool)
    return data[~stopped_mask].copy(), data[stopped_mask].copy()


def normalize_rate(series: pd.Series) -> pd.Series:
    numeric = normalize_config_number(series, percent_to_decimal=True)
    return numeric.where(numeric <= 1, numeric / 100)


def normalize_config_number(series: pd.Series, percent_to_decimal: bool = False) -> pd.Series:
    text = normalized_numeric_text(series)
    percent_mask = text.str.endswith("%", na=False)
    cleaned = text.str.replace("%", "", regex=False)
    numeric = pd.to_numeric(cleaned, errors="coerce").astype("float64")
    if percent_to_decimal:
        numeric.loc[percent_mask] = numeric.loc[percent_mask] / 100
    return numeric


def normalized_numeric_text(series: pd.Series) -> pd.Series:
    return series.map(clean_numeric_text)


def clean_numeric_text(value) -> str | None:
    if pd.isna(value):
        return None
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if text.lower() in {"", "nan", "none", "null", "nat", "-", "--", "—", "–", "未配置"}:
        return None

    is_parenthesized_negative = text.startswith("(") and text.endswith(")")
    if is_parenthesized_negative:
        text = text[1:-1].strip()

    text = (
        text.replace("\u00a0", "")
        .replace("\u2007", "")
        .replace("\u202f", "")
        .replace("\u3000", "")
        .replace(" ", "")
        .replace("\t", "")
        .replace(",", "")
        .replace("，", "")
        .replace("￥", "")
        .replace("¥", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
    )
    text = text.replace("−", "-").replace("–", "-").replace("—", "-")
    if is_parenthesized_negative and text and not text.startswith("-"):
        text = "-" + text
    if not re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)%?", text):
        match = re.match(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)%?", text)
        if match:
            text = match.group(0)
    return text


def build_discovered_commission_config(reports: pd.DataFrame | None) -> pd.DataFrame:
    if reports is None or reports.empty:
        return normalize_commission_config(pd.DataFrame()).fillna("")
    commission = (
        reports[["月份", "销售专员"]]
        .dropna()
        .drop_duplicates()
        .rename(columns={"销售专员": "开发员"})
        .sort_values(["月份", "开发员"])
    )
    commission["库存计提"] = ""
    commission["弃置"] = ""
    commission["职位提点"] = ""
    return normalize_commission_config(commission).fillna("")


def build_discovered_department_fee_config(reports: pd.DataFrame | None) -> pd.DataFrame:
    if reports is None or reports.empty or "部门" not in reports.columns:
        return normalize_department_fee_config(pd.DataFrame()).fillna("")
    fee_config = (
        reports[["月份", "部门"]]
        .dropna()
        .drop_duplicates()
        .sort_values(["月份", "部门"])
    )
    fee_config["费用率"] = ""
    return normalize_department_fee_config(fee_config).fillna("")


def select_metric_config(metric_config: pd.DataFrame, metric_names: list[str]) -> pd.DataFrame:
    if metric_config.empty:
        raise ValueError("指标配置为空，无法计算提成")
    group_rank = {"开发员分析": 0, "总览": 1, "全部": 2, "趋势": 3, "店铺分析": 4, "开发员店铺分析": 5}
    selected = metric_config[metric_config["指标名称"].isin(metric_names)].copy()
    selected["_metric_order"] = selected["指标名称"].map({name: idx for idx, name in enumerate(metric_names)})
    selected["_group_rank"] = selected["显示分组"].map(group_rank).fillna(99)
    selected = selected.sort_values(["_metric_order", "_group_rank"]).drop_duplicates("指标名称", keep="first")
    missing = [name for name in metric_names if name not in set(selected["指标名称"])]
    if missing:
        raise ValueError(f"提成计算缺少指标公式：{', '.join(missing)}")
    return selected.drop(columns=["_metric_order", "_group_rank"])


def compute_commission_table(
    df: pd.DataFrame,
    metric_config: pd.DataFrame,
    commission_config: pd.DataFrame,
    department_fee_config: pd.DataFrame,
) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=COMMISSION_OUTPUT_COLUMNS)

    metrics = select_metric_config(metric_config, ["销售额", "毛利润", "毛利率"])
    base = compute_metric_table(df, metrics, ["月份", "销售专员", "部门"]).rename(
        columns={"销售专员": "开发员", "销售额": "营业额"}
    )
    if base.empty:
        return pd.DataFrame(columns=COMMISSION_OUTPUT_COLUMNS)

    fee_config = normalize_department_fee_config(department_fee_config).copy()
    fee_config["__has_department_fee_config"] = True
    base = base.merge(fee_config, on=["月份", "部门"], how="left")
    if "__has_department_fee_config" not in base.columns:
        base["__has_department_fee_config"] = False
    base["__has_department_fee_config"] = base["__has_department_fee_config"].fillna(False)
    base["__部门费用额"] = base["营业额"] * base["费用率"]
    base["__部门提成前利润"] = base["营业额"] * (base["毛利率"] - base["费用率"])

    has_department_fee = base["__has_department_fee_config"].fillna(False).astype(bool) & base["费用率"].notna()
    dept_missing = (
        base[~has_department_fee]
        .groupby(["月份", "开发员"], dropna=False)
        .size()
        .rename("__缺部门费用率数")
        .reset_index()
    )
    base_summary = (
        base.groupby(["月份", "开发员"], dropna=False, as_index=False)
        .agg(
            营业额=("营业额", "sum"),
            毛利润=("毛利润", "sum"),
            __部门费用额=("__部门费用额", "sum"),
            __部门提成前利润=("__部门提成前利润", "sum"),
        )
        .merge(dept_missing, on=["月份", "开发员"], how="left")
    )
    base_summary["__缺部门费用率数"] = base_summary["__缺部门费用率数"].fillna(0)
    base_summary["毛利率"] = base_summary.apply(
        lambda row: row["毛利润"] / row["营业额"] if pd.notna(row["营业额"]) and row["营业额"] else pd.NA,
        axis=1,
    )
    base_summary["费用率"] = base_summary.apply(
        lambda row: row["__部门费用额"] / row["营业额"] if pd.notna(row["营业额"]) and row["营业额"] else pd.NA,
        axis=1,
    )

    config = normalize_commission_config(commission_config).copy()
    config["__has_commission_config"] = True
    merged = base_summary.merge(config, on=["月份", "开发员"], how="left")
    if "__has_commission_config" not in merged.columns:
        merged["__has_commission_config"] = False
    merged["__has_commission_config"] = merged["__has_commission_config"].fillna(False)

    param_cols = ["库存计提", "弃置", "职位提点"]
    has_all_params = (
        merged["__has_commission_config"]
        & merged[param_cols].notna().all(axis=1)
        & merged["__缺部门费用率数"].eq(0)
    )
    merged["配置状态"] = has_all_params.map(lambda value: "已配置" if value else "缺配置")
    merged["提成预估"] = pd.NA
    merged.loc[has_all_params, "提成预估"] = (
        (merged.loc[has_all_params, "__部门提成前利润"] - merged.loc[has_all_params, "库存计提"] - merged.loc[has_all_params, "弃置"])
        * merged.loc[has_all_params, "职位提点"]
    )
    return merged[COMMISSION_OUTPUT_COLUMNS].sort_values(["月份", "开发员"]).reset_index(drop=True)


def build_person_commission_summary(
    df: pd.DataFrame,
    metric_config: pd.DataFrame,
    commission_config: pd.DataFrame,
    department_fee_config: pd.DataFrame,
) -> pd.DataFrame:
    columns = ["人员", "营业额", "毛利润", "毛利率", "提成预估", "缺配置月份数"]
    detail = compute_commission_table(df, metric_config, commission_config, department_fee_config)
    if detail.empty:
        return pd.DataFrame(columns=columns)

    numeric = detail.copy()
    for col in ["营业额", "毛利润", "提成预估"]:
        numeric[col] = pd.to_numeric(numeric[col], errors="coerce")
    numeric["人员"] = numeric["开发员"].map(normalize_department_person_name)
    numeric["人员"] = numeric["人员"].where(numeric["人员"].astype(str).str.strip().ne(""), numeric["开发员"])
    numeric["_缺配置"] = numeric["配置状态"].ne("已配置")

    summary = (
        numeric.groupby("人员", dropna=False, as_index=False)
        .agg(
            营业额=("营业额", "sum"),
            毛利润=("毛利润", "sum"),
            提成预估=("提成预估", lambda values: values.sum(min_count=1)),
            缺配置月份数=("_缺配置", "sum"),
        )
    )
    summary["毛利率"] = summary.apply(
        lambda row: row["毛利润"] / row["营业额"] if pd.notna(row["营业额"]) and row["营业额"] else pd.NA,
        axis=1,
    )
    summary = summary[columns].sort_values("提成预估", ascending=False, na_position="last").reset_index(drop=True)

    total_row = {
        "人员": "合计",
        "营业额": summary["营业额"].sum(),
        "毛利润": summary["毛利润"].sum(),
        "提成预估": summary["提成预估"].sum(min_count=1),
        "缺配置月份数": summary["缺配置月份数"].sum(),
    }
    total_row["毛利率"] = (
        total_row["毛利润"] / total_row["营业额"] if pd.notna(total_row["营业额"]) and total_row["营业额"] else pd.NA
    )
    return pd.concat([summary, pd.DataFrame([total_row])[columns]], ignore_index=True)


def compute_stopped_commission_table(
    df: pd.DataFrame,
    metric_config: pd.DataFrame,
    commission_config: pd.DataFrame,
    department_fee_config: pd.DataFrame,
) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=STOPPED_COMMISSION_OUTPUT_COLUMNS)

    metrics = select_metric_config(metric_config, ["销售额", "毛利润", "毛利率"])
    group_cols = ["月份", "销售专员", "店铺编码", "店铺类型", "部门", "停提款时间"]
    base = compute_metric_table(df, metrics, group_cols).rename(
        columns={"销售专员": "开发员", "销售额": "营业额"}
    )
    if base.empty:
        return pd.DataFrame(columns=STOPPED_COMMISSION_OUTPUT_COLUMNS)

    fee_config = normalize_department_fee_config(department_fee_config).copy()
    fee_config["__has_department_fee_config"] = True
    base = base.merge(fee_config, on=["月份", "部门"], how="left")
    if "__has_department_fee_config" not in base.columns:
        base["__has_department_fee_config"] = False
    base["__has_department_fee_config"] = base["__has_department_fee_config"].fillna(False)

    config = normalize_commission_config(commission_config).copy()
    config["__has_commission_config"] = True
    merged = base.merge(config, on=["月份", "开发员"], how="left")
    if "__has_commission_config" not in merged.columns:
        merged["__has_commission_config"] = False
    merged["__has_commission_config"] = merged["__has_commission_config"].fillna(False)

    merged["__月开发员停提款营业额"] = merged.groupby(["月份", "开发员"])["营业额"].transform("sum")
    merged["__分摊比例"] = merged.apply(
        lambda row: row["营业额"] / row["__月开发员停提款营业额"] if row["__月开发员停提款营业额"] else 0,
        axis=1,
    )
    merged["库存计提分摊"] = merged["库存计提"] * merged["__分摊比例"]
    merged["弃置分摊"] = merged["弃置"] * merged["__分摊比例"]

    param_cols = ["库存计提", "弃置", "职位提点"]
    has_all_params = (
        merged["__has_commission_config"]
        & merged[param_cols].notna().all(axis=1)
        & merged["__has_department_fee_config"].fillna(False).astype(bool)
        & merged["费用率"].notna()
    )
    merged["配置状态"] = has_all_params.map(lambda value: "已配置" if value else "缺配置")
    merged["缺提成预估"] = pd.NA
    merged.loc[has_all_params, "缺提成预估"] = (
        (
            merged.loc[has_all_params, "营业额"] * (merged.loc[has_all_params, "毛利率"] - merged.loc[has_all_params, "费用率"])
            - merged.loc[has_all_params, "库存计提分摊"]
            - merged.loc[has_all_params, "弃置分摊"]
        )
        * merged.loc[has_all_params, "职位提点"]
    )
    return merged[STOPPED_COMMISSION_OUTPUT_COLUMNS].sort_values(["月份", "开发员", "店铺编码"]).reset_index(drop=True)


def compute_metric_table(df: pd.DataFrame, metric_config: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if metric_config.empty:
        return pd.DataFrame()

    validate_metric_fields(df, metric_config)
    working = df.reset_index(drop=True)

    field_names = {
        field
        for formula in metric_config["公式"]
        for field in extract_fields(formula)
        if field in working.columns
    }
    field_cache = {field: maybe_numeric(working[field]) for field in field_names}
    range_specs = {
        bounds
        for formula in metric_config["公式"]
        for bounds in extract_range_sums(formula)
    }
    range_sum_cache: dict[tuple[str, str], pd.Series] = {}
    columns = list(working.columns)
    for start, end in range_specs:
        if start not in columns or end not in columns:
            continue
        start_idx, end_idx = columns.index(start), columns.index(end)
        if start_idx > end_idx:
            continue
        numeric_columns = {
            column: normalize_config_number(working[column])
            for column in columns[start_idx : end_idx + 1]
        }
        range_sum_cache[(start, end)] = pd.DataFrame(numeric_columns, index=working.index).sum(axis=1)

    target_month_count = 1
    if "月份" in working.columns and "月份" not in group_cols:
        target_month_count = max(1, int(working["月份"].dropna().astype(str).nunique()))

    rows = []
    if group_cols:
        grouped = working.groupby(group_cols, dropna=False, sort=False)
        for keys, group in grouped:
            if not isinstance(keys, tuple):
                keys = (keys,)
            row = dict(zip(group_cols, keys))
            row.update(
                compute_metrics_for_frame(
                    group,
                    metric_config,
                    target_month_count=target_month_count,
                    field_cache=field_cache,
                    range_sum_cache=range_sum_cache,
                )
            )
            rows.append(row)
    else:
        rows.append(
            compute_metrics_for_frame(
                working,
                metric_config,
                target_month_count=target_month_count,
                field_cache=field_cache,
                range_sum_cache=range_sum_cache,
            )
        )
    return pd.DataFrame(rows)


def validate_metric_fields(df: pd.DataFrame, metric_config: pd.DataFrame) -> None:
    missing_by_metric = []
    columns = set(df.columns)
    for _, metric in metric_config.iterrows():
        missing = [field for field in extract_fields(metric["公式"]) if field not in columns]
        if missing:
            missing_by_metric.append(f"{metric['指标名称']} 缺少字段：{', '.join(missing)}")
    if missing_by_metric:
        raise FormulaError("; ".join(missing_by_metric))


def compute_metrics_for_frame(
    df: pd.DataFrame,
    metric_config: pd.DataFrame,
    target_month_count: int = 1,
    field_cache: dict[str, pd.Series] | None = None,
    range_sum_cache: dict[tuple[str, str], pd.Series] | None = None,
) -> dict:
    def get_field(name: str):
        if name not in df.columns:
            raise FormulaError(f"字段不存在：{name}")
        values = field_cache[name].loc[df.index] if field_cache is not None and name in field_cache else maybe_numeric(df[name])
        if name == "销售额目标" and target_month_count > 1:
            return values * target_month_count
        return values

    def get_range_sum(start: str, end: str):
        cached = range_sum_cache.get((start, end)) if range_sum_cache is not None else None
        if cached is not None:
            return cached.loc[df.index].sum()
        columns = list(df.columns)
        if start not in columns:
            raise FormulaError(f"range_sum() 起始字段不存在：{start}")
        if end not in columns:
            raise FormulaError(f"range_sum() 结束字段不存在：{end}")
        start_idx = columns.index(start)
        end_idx = columns.index(end)
        if start_idx > end_idx:
            raise FormulaError(f"range_sum() 起始字段不能在结束字段之后：{start} > {end}")
        numeric_range = df.loc[:, columns[start_idx : end_idx + 1]].apply(normalize_config_number)
        return numeric_range.sum().sum()

    context = FormulaContext(field_getter=get_field, range_sum_getter=get_range_sum)
    row = {}
    for _, metric in metric_config.iterrows():
        value = evaluate_formula(metric["公式"], context)
        if isinstance(value, pd.Series):
            value = normalize_config_number(value).sum()
        row[metric["指标名称"]] = value
    return row


def format_display_table(df: pd.DataFrame, metric_lookup: dict) -> pd.DataFrame:
    result = df.copy()
    for col in result.columns:
        fmt = metric_lookup.get(col, {}).get("格式")
        if not fmt:
            continue
        if fmt == "金额":
            result[col] = result[col].map(lambda value: format_number(value, ",.2f"))
        elif fmt == "整数":
            result[col] = result[col].map(lambda value: format_number(value, ",.0f"))
        elif fmt == "百分比":
            result[col] = result[col].map(lambda value: "-" if pd.isna(value) else f"{float(value):.2%}")
        else:
            result[col] = result[col].map(lambda value: format_number(value, ",.2f"))
    return result


def format_number(value, pattern: str) -> str:
    if pd.isna(value):
        return "-"
    try:
        return format(float(value), pattern)
    except (TypeError, ValueError):
        return str(value)


def build_alerts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    alert_rows = []
    for _, row in df.iterrows():
        reasons = []
        if "毛利率" in row and pd.notna(row["毛利率"]) and row["毛利率"] < 0.15:
            reasons.append("毛利率低于 15%")
        if "广告费占比" in row and pd.notna(row["广告费占比"]) and row["广告费占比"] > 0.1:
            reasons.append("广告费占比高于 10%")
        if "目标完成率" in row and pd.notna(row["目标完成率"]) and row["目标完成率"] < 0.9:
            reasons.append("目标完成率低于 90%")
        if reasons:
            item = row.to_dict()
            item["预警原因"] = "；".join(reasons)
            alert_rows.append(item)
    return pd.DataFrame(alert_rows)
